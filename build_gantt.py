from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
EXCEL_FILE = ROOT / "Tidplan.xlsx"
OUT_MD = ROOT / "tidplan_gantt_generated.md"
OUT_STATUS_MD = ROOT / "tidplan_gantt_status_per_land.md"
OUT_STATUS_HTML = ROOT / "tidplan_gantt_status_per_land.html"
OUT_CALC_HTML = ROOT / "tidplan_gantt_kalkyl.html"
OUT_INDEX_HTML = ROOT / "index.html"
BACKUP_STATUS_HTML = ROOT / "backups" / "latest_snapshot" / "tidplan_gantt_status_per_land.html"
MASTER_BACKUP_FILE = ROOT / "backups" / "MASTER_BACKUP.txt"

DRAG_EDITOR_REQUIRED_SNIPPETS = (
    '<h2 class="editor-title">Drag editor (drag name or bar)</h2>',
    '<div id="dragEditor" class="editor"></div>',
    "const dragEditor = document.getElementById('dragEditor');",
    "function renderDragEditor(visibleRows, fromDate, toDate)",
    "function addDragToReschedule(visibleRows)",
)


def iso(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def slug(s: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in s).strip("_")


def country_short(country: str) -> str:
    mapping = {
        "sweden": "swe",
        "norway": "nor",
        "finland": "fin",
        "lithuania": "ltu",
    }
    key = (country or "").strip().lower()
    if key in mapping:
        return mapping[key]
    fallback = "".join(ch for ch in key if ch.isalpha())
    return (fallback[:3] or "n/a").lower()


def normalize_status(value: str | None) -> str:
    raw = (value or "").strip().lower()
    if raw in {"done", "klar", "gron", "green", "complete", "completed", "100", "full"}:
        return "done"
    if raw in {"partial", "delvis", "half", "50", "halv"}:
        return "partial"
    return "none"


def find_header_row(ws):
    for r in range(1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if (
            "Priority" in values
            and "Country" in values
            and "Server" in values
            and "Upgrade Date" in values
        ):
            return r, values
    raise RuntimeError("Could not find header row with Priority/Country/Server/Upgrade Date")


def parse_rows():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row, headers = find_header_row(ws)
    idx = {name: headers.index(name) + 1 for name in ["Priority", "Country", "Server", "Environment", "Upgrade Date"]}
    status_col = headers.index("Status") + 1 if "Status" in headers else None
    rhel_ok_col = None
    for candidate in ("RHEL9 OK", "RHEL OK", "RHEL9", "RHEL"):
        if candidate in headers:
            rhel_ok_col = headers.index(candidate) + 1
            break
    ime_ok_col = None
    for candidate in ("IME V11 OK", "IME OK", "IME V11", "IME"):
        if candidate in headers:
            ime_ok_col = headers.index(candidate) + 1
            break
    app_test_ok_col = None
    for candidate in ("Application Test OK", "Applikationstest OK", "App Test OK", "Applicationstest OK"):
        if candidate in headers:
            app_test_ok_col = headers.index(candidate) + 1
            break
    regression_test_ok_col = None
    for candidate in ("Regression Test OK", "Regressionstest OK", "Regression OK"):
        if candidate in headers:
            regression_test_ok_col = headers.index(candidate) + 1
            break
    ip_swap_ok_col = None
    for candidate in ("IP Swap OK", "IP Swap Klar", "IP Swap Done"):
        if candidate in headers:
            ip_swap_ok_col = headers.index(candidate) + 1
            break
    ip_swap_start_date_col = None
    for candidate in ("IP Swap From Date", "IP Swap Start Date", "IP Swap Date", "IP Swap Datum", "IP Swap Planned Date"):
        if candidate in headers:
            ip_swap_start_date_col = headers.index(candidate) + 1
            break
    ip_swap_end_date_col = None
    for candidate in ("IP Swap Tom Date", "IP Swap To Date", "IP Swap End Date", "IP Swap Slutdatum"):
        if candidate in headers:
            ip_swap_end_date_col = headers.index(candidate) + 1
            break
    ip_swap_from_col = None
    for candidate in ("IP Swap From", "IP Swap Fran", "IP Swap Från", "Swap From Server"):
        if candidate in headers:
            ip_swap_from_col = headers.index(candidate) + 1
            break
    ip_swap_to_col = None
    for candidate in ("IP Swap To", "Swap To Server", "IP Swap Till"):
        if candidate in headers:
            ip_swap_to_col = headers.index(candidate) + 1
            break
    patch_ok_col = None
    for candidate in ("Patchning OK", "Patch OK", "Patch Done", "Patch"):
        if candidate in headers:
            patch_ok_col = headers.index(candidate) + 1
            break
    patch_number_col = None
    for candidate in ("Patchnummer", "Patch Number", "Patch Version"):
        if candidate in headers:
            patch_number_col = headers.index(candidate) + 1
            break
    notes_prod_col = None
    for candidate in ("Notes Prod", "Prod Notes", "Anteckning Prod", "Prod Anteckning"):
        if candidate in headers:
            notes_prod_col = headers.index(candidate) + 1
            break
    notes_test_col = None
    for candidate in ("Notes Test", "Test Notes", "Anteckning Test", "Test Anteckning"):
        if candidate in headers:
            notes_test_col = headers.index(candidate) + 1
            break
    notes_col = None
    for candidate in ("Notes", "Note", "Kommentar", "Anteckning"):
        if candidate in headers:
            notes_col = headers.index(candidate) + 1
            break

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        priority = ws.cell(r, idx["Priority"]).value
        country = ws.cell(r, idx["Country"]).value
        server = ws.cell(r, idx["Server"]).value
        environment = ws.cell(r, idx["Environment"]).value
        upgrade_date = ws.cell(r, idx["Upgrade Date"]).value
        status_value = ws.cell(r, status_col).value if status_col else None
        rhel_ok_value = ws.cell(r, rhel_ok_col).value if rhel_ok_col else None
        ime_ok_value = ws.cell(r, ime_ok_col).value if ime_ok_col else None
        app_test_ok_value = ws.cell(r, app_test_ok_col).value if app_test_ok_col else None
        regression_test_ok_value = ws.cell(r, regression_test_ok_col).value if regression_test_ok_col else None
        ip_swap_ok_value = ws.cell(r, ip_swap_ok_col).value if ip_swap_ok_col else None
        ip_swap_start_date_value = ws.cell(r, ip_swap_start_date_col).value if ip_swap_start_date_col else None
        ip_swap_end_date_value = ws.cell(r, ip_swap_end_date_col).value if ip_swap_end_date_col else None
        ip_swap_from_value = ws.cell(r, ip_swap_from_col).value if ip_swap_from_col else None
        ip_swap_to_value = ws.cell(r, ip_swap_to_col).value if ip_swap_to_col else None
        patch_ok_value = ws.cell(r, patch_ok_col).value if patch_ok_col else None
        patch_number_value = ws.cell(r, patch_number_col).value if patch_number_col else None
        notes_prod_value = ws.cell(r, notes_prod_col).value if notes_prod_col else None
        notes_test_value = ws.cell(r, notes_test_col).value if notes_test_col else None
        notes_value = ws.cell(r, notes_col).value if notes_col else None

        if not (priority and country and server and upgrade_date):
            continue

        if isinstance(upgrade_date, datetime):
            upg_date = upgrade_date.date()
        elif isinstance(upgrade_date, date):
            upg_date = upgrade_date
        else:
            upg_date = datetime.strptime(str(upgrade_date), "%Y-%m-%d").date()

        rows.append(
            {
                "priority": int(priority),
                "country": str(country).strip(),
                "server": str(server).strip(),
                "environment": str(environment).strip() if environment else "",
                "date": upg_date,
                "duration": 1,
                "status": normalize_status(str(status_value) if status_value is not None else None),
                "rhel_ok": str(rhel_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if rhel_ok_value is not None
                else False,
                "ime_ok": str(ime_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if ime_ok_value is not None
                else False,
                "app_test_ok": str(app_test_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if app_test_ok_value is not None
                else False,
                "regression_test_ok": str(regression_test_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if regression_test_ok_value is not None
                else False,
                "ip_swap_ok": str(ip_swap_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if ip_swap_ok_value is not None
                else False,
                "ip_swap_start_date": iso(ip_swap_start_date_value.date())
                if isinstance(ip_swap_start_date_value, datetime)
                else iso(ip_swap_start_date_value)
                if isinstance(ip_swap_start_date_value, date)
                else str(ip_swap_start_date_value).strip()
                if ip_swap_start_date_value is not None
                else "",
                "ip_swap_end_date": iso(ip_swap_end_date_value.date())
                if isinstance(ip_swap_end_date_value, datetime)
                else iso(ip_swap_end_date_value)
                if isinstance(ip_swap_end_date_value, date)
                else str(ip_swap_end_date_value).strip()
                if ip_swap_end_date_value is not None
                else (
                    iso(ip_swap_start_date_value.date())
                    if isinstance(ip_swap_start_date_value, datetime)
                    else iso(ip_swap_start_date_value)
                    if isinstance(ip_swap_start_date_value, date)
                    else str(ip_swap_start_date_value).strip()
                    if ip_swap_start_date_value is not None
                    else ""
                ),
                "ip_swap_from": str(ip_swap_from_value).strip() if ip_swap_from_value is not None else "",
                "ip_swap_to": str(ip_swap_to_value).strip() if ip_swap_to_value is not None else "",
                "patch_ok": str(patch_ok_value).strip().lower() in {"1", "true", "yes", "ja", "y", "ok", "done", "klar"}
                if patch_ok_value is not None
                else False,
                "patch_number": str(patch_number_value).strip() if patch_number_value is not None else "",
                "notes_prod": str(notes_prod_value).strip() if notes_prod_value is not None else (str(notes_value).strip() if notes_value is not None else ""),
                "notes_test": str(notes_test_value).strip() if notes_test_value is not None else "",
                "notes": str(notes_value).strip() if notes_value is not None else "",
            }
        )

    if not rows:
        raise RuntimeError("No plan rows found in Tidplan.xlsx")

    rows.sort(key=lambda x: (x["priority"], x["date"], x["server"]))
    return rows


def country_order(rows):
    order = []
    for row in rows:
        country = row["country"]
        if country not in order:
            order.append(country)
    return order


def read_master_backup_label() -> str:
    if not MASTER_BACKUP_FILE.exists():
        return "No master copy saved"
    lines = [line.strip() for line in MASTER_BACKUP_FILE.read_text(encoding="utf-8").splitlines() if line.strip()]
    return lines[0] if lines else "No master copy saved"


def build_mermaid(rows, countries, title):
    first_date = min(r["date"] for r in rows)
    last_date = max(r["date"] for r in rows)
    frame_start = date(first_date.year, first_date.month, 1)

    plan_days = (last_date - frame_start).days + 1

    lines = [
        "gantt",
        f"    title {title}",
        "    dateFormat  YYYY-MM-DD",
        "    axisFormat  %Y-%m-%d",
        "",
        "    section Timeline",
        f"    Start date                                  :milestone, frame_start, {iso(frame_start)}, 1d",
        f"    Full plan period                            :frame_period, {iso(frame_start)}, {plan_days}d",
        f"    End date                                    :milestone, frame_end, {iso(last_date)}, 1d",
        "",
    ]

    for country in countries:
        lines.append(f"    section {country}")
        cslug = slug(country)
        country_rows = [r for r in rows if r["country"] == country]
        for i, row in enumerate(country_rows, start=1):
            base = f"{row['server']} ({row['environment']})" if row["environment"] else row["server"]
            label = f"{base} ({country_short(row['country'])})"
            duration = max(1, int(row.get("duration", 1)))
            lines.append(f"    {label:<45} :{cslug}_{i}, {iso(row['date'])}, {duration}d")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n", frame_start, last_date


def build_html(rows, countries, frame_start, last_date, planner_mode=False):
    button_items = ['        <button class="country-btn active" data-country="all">All countries</button>']
    for country in countries:
        button_items.append(f'        <button class="country-btn" data-country="{country}">{country}</button>')

    rows_payload = [
        {
            "priority": int(row.get("priority", 9999)),
            "country": row["country"],
            "server": row["server"],
            "environment": row["environment"],
            "date": iso(row["date"]),
            "duration": int(row.get("duration", 1)),
            "status": row.get("status", "none"),
            "rhelOk": bool(row.get("rhel_ok", False)),
            "imeOk": bool(row.get("ime_ok", False)),
            "appTestOk": bool(row.get("app_test_ok", False)),
            "regressionTestOk": bool(row.get("regression_test_ok", False)),
            "ipSwapOk": bool(row.get("ip_swap_ok", False)),
            "ipSwapStartDate": str(row.get("ip_swap_start_date", "")),
            "ipSwapEndDate": str(row.get("ip_swap_end_date", row.get("ip_swap_start_date", ""))),
            "ipSwapFrom": str(row.get("ip_swap_from", "")),
            "ipSwapTo": str(row.get("ip_swap_to", "")),
            "patchOk": bool(row.get("patch_ok", False)),
            "patchNumber": str(row.get("patch_number", "")),
            "notesProd": str(row.get("notes_prod", row.get("notes", ""))),
            "notesTest": str(row.get("notes_test", "")),
            "notes": str(row.get("notes", "")),
        }
        for row in rows
    ]

    countries_json = json.dumps(countries, ensure_ascii=True)
    rows_json = json.dumps(rows_payload, ensure_ascii=True)
    frame_start_iso = iso(frame_start)
    last_date_iso = iso(last_date)
    master_backup_label = read_master_backup_label().replace("'", "&#39;")
    title_text = "Upgrade Plan 2026 - Planning View" if planner_mode else "Upgrade Plan 2026 - Production Schedule"
    switch_link = (
        '                                <a class="btn-secondary nav-link" href="tidplan_gantt_status_per_land.html">Go to production view</a>'
        if planner_mode
        else '                                <a class="btn-secondary nav-link" href="tidplan_gantt_kalkyl.html">Go to planning view</a>'
    )
    master_copy_file = "tidplan_gantt_kalkyl.html" if planner_mode else "tidplan_gantt_status_per_land.html"
    planner_section = (
        f'''            <div class="planner">
                <div class="planner-title">Planning suggestion</div>
                <div class="field">
                    <label for="planStartDate">Plan start</label>
                    <input id="planStartDate" type="date" value="{frame_start_iso}" />
                </div>
                <div class="field">
                    <label for="planPerDay">Servers per day</label>
                    <input id="planPerDay" type="number" value="2" min="1" max="20" step="1" />
                </div>
                <label class="field-check">
                    <input id="planSkipWeekends" type="checkbox" checked />
                    Skip weekends
                </label>
                <button id="planSuggestBtn" class="btn-secondary" type="button" onclick="window.__planSuggest && window.__planSuggest()">Calculate suggestion</button>
            </div>'''
        if planner_mode
        else ""
    )

    return f'''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>Gantt by country</title>
  <style>
    body {{ font-family: Segoe UI, Tahoma, sans-serif; margin: 0; background: #f4f7fb; color: #1f2937; }}
    .wrap {{ max-width: 1400px; margin: 20px auto; padding: 0 16px; }}
    .card {{ background: #fff; border: 1px solid #d1d5db; border-radius: 10px; padding: 16px; }}
    .page-head {{ display: flex; justify-content: space-between; align-items: center; gap: 10px; margin-bottom: 8px; }}
    .head-actions {{ display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }}
    h1 {{ margin-top: 0; }}
    .meta {{ color: #4b5563; margin-bottom: 12px; }}
        .overview {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 8px;
            margin: 0 0 12px;
        }}
        .overview-card {{
            border: 1px solid #dbe2ea;
            border-radius: 10px;
            padding: 8px 10px;
            background: #f8fafc;
            display: grid;
            gap: 2px;
            cursor: pointer;
            transition: transform 120ms ease, box-shadow 120ms ease, border-color 120ms ease;
        }}
        .overview-card:hover {{
            transform: translateY(-1px);
            border-color: #93c5fd;
            box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
        }}
        .overview-card:focus-visible {{
            outline: 2px solid #60a5fa;
            outline-offset: 2px;
        }}
        .overview-label {{
            color: #64748b;
            font-size: 0.78rem;
            letter-spacing: 0.01em;
            text-transform: uppercase;
        }}
        .overview-value {{
            color: #0f172a;
            font-size: 1.2rem;
            font-weight: 700;
            line-height: 1.1;
        }}
        .overview-card.done .overview-value {{ color: #15803d; }}
        .overview-card.partial .overview-value {{ color: #b45309; }}
        .overview-card.remaining .overview-value {{ color: #b91c1c; }}
        .controls {{ display: flex; gap: 10px; flex-wrap: wrap; margin: 8px 0 12px; }}
        .country-btn {{
            border: 1px solid #cbd5e1;
            background: #f8fafc;
            color: #0f172a;
            border-radius: 999px;
            padding: 6px 12px;
            font-size: 0.92rem;
            cursor: pointer;
        }}
        .country-btn.active {{ background: #dbeafe; border-color: #93c5fd; }}
        .filters {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            align-items: end;
            margin: 0 0 14px;
            padding: 10px;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #f8fafc;
        }}
        .planner {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            align-items: end;
            margin: 0 0 14px;
            padding: 10px;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #eef6ff;
            position: relative;
            z-index: 5;
        }}
        .planner button,
        .planner input,
        .planner label {{ pointer-events: auto; }}
        .planner-title {{
            width: 100%;
            font-size: 0.9rem;
            font-weight: 600;
            color: #1e3a8a;
            margin-bottom: -2px;
        }}
        .field-check {{
            display: flex;
            gap: 6px;
            align-items: center;
            color: #334155;
            font-size: 0.86rem;
            padding-bottom: 4px;
        }}
        .field {{ display: grid; gap: 4px; }}
        .field label {{ font-size: 0.84rem; color: #475569; }}
        .field input {{
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 6px 8px;
            font: inherit;
            color: #0f172a;
            background: #fff;
        }}
        .btn-secondary {{
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 6px 10px;
            background: #fff;
            color: #0f172a;
            cursor: pointer;
            font: inherit;
        }}
        .nav-link {{ text-decoration: none; display: inline-block; }}
        .btn-fullscreen {{ white-space: nowrap; }}
        .status {{ margin-bottom: 10px; color: #334155; font-size: 0.92rem; }}
        .error {{ color: #b91c1c; }}
        .meta-row {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 12px;
            flex-wrap: wrap;
            margin-bottom: 12px;
        }}
        .meta-row .meta {{ margin-bottom: 0; }}
        .master-tools {{ display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }}
        .master-label {{ color: #475569; font-size: 0.88rem; }}
        .overview-modal-backdrop {{
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.45);
            display: none;
            align-items: center;
            justify-content: center;
            z-index: 10030;
            padding: 16px;
        }}
        .overview-modal-backdrop.open {{ display: flex; }}
        .overview-modal {{
            width: min(760px, 96vw);
            max-height: min(78vh, 760px);
            overflow: hidden;
            background: #ffffff;
            border: 1px solid #cbd5e1;
            border-radius: 12px;
            box-shadow: 0 18px 44px rgba(15, 23, 42, 0.28);
            padding: 12px;
            display: grid;
            gap: 10px;
        }}
        .overview-modal h3 {{
            margin: 0;
            font-size: 1rem;
            color: #0f172a;
        }}
        .overview-modal-list {{
            margin: 0;
            padding-left: 18px;
            overflow: auto;
            display: grid;
            gap: 6px;
            color: #1f2937;
        }}
        .overview-modal-list li {{
            font-size: 0.9rem;
            line-height: 1.3;
        }}
        .overview-modal-actions {{
            display: flex;
            justify-content: flex-end;
            gap: 8px;
            flex-wrap: wrap;
        }}
        .overview-empty {{
            color: #64748b;
            font-size: 0.9rem;
            margin: 0;
        }}
    .mermaid {{ overflow: auto; }}
        #chart {{ min-height: 120px; }}
        .editor-title {{ margin: 14px 0 8px; font-size: 1rem; color: #0f172a; }}
        .editor-filters {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
            gap: 8px;
            margin: 0 0 10px;
            align-items: end;
        }}
        .editor-filters .field {{ margin: 0; }}
        .editor-filters .field select,
        .editor-filters .field input {{
            width: 100%;
            height: 36px;
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 6px 10px;
            font: inherit;
            color: #0f172a;
            background: #fff;
            box-sizing: border-box;
        }}
        #dragFilterReset {{
            height: 36px;
            align-self: end;
            white-space: nowrap;
        }}
        .editor {{
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #f8fafc;
            padding: 10px;
            display: grid;
            gap: 10px;
            overflow-x: auto;
        }}
        .editor-row {{
            display: grid;
            grid-template-columns: minmax(240px, 1.1fr) minmax(420px, 2.9fr);
            gap: 12px;
            align-items: center;
            min-width: 720px;
        }}
        .editor-name {{
            font-size: 0.92rem;
            color: #0f172a;
            cursor: grab;
            user-select: none;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .editor-name-wrap {{
            display: flex;
            align-items: center;
            gap: 8px;
            min-width: 0;
        }}
        .note-badge {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-width: 18px;
            height: 18px;
            padding: 0 5px;
            border-radius: 999px;
            border: 1px solid #0284c7;
            color: #075985;
            background: #e0f2fe;
            font-size: 0.72rem;
            font-weight: 700;
            line-height: 1;
            flex-shrink: 0;
        }}
        .editor-track {{
            position: relative;
            height: 26px;
            border: 1px dashed #cbd5e1;
            border-radius: 8px;
            background: #fff;
            overflow: visible;
        }}
        .editor-bar {{
            position: absolute;
            top: 4px;
            height: 16px;
            min-width: 12px;
            border-radius: 6px;
            background: #0ea5e9;
            cursor: grab;
            z-index: 2;
        }}
        .editor-live {{
            position: absolute;
            top: -24px;
            left: 0;
            display: none;
            background: #0f172a;
            color: #fff;
            font-size: 0.75rem;
            border-radius: 6px;
            padding: 2px 6px;
            white-space: nowrap;
            pointer-events: none;
            z-index: 6;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.2);
        }}
        .editor-live.open {{ display: inline-block; }}
        .editor-mini {{ border: 1px solid #cbd5e1; border-radius: 6px; background: #fff; padding: 3px 8px; cursor: pointer; font: inherit; line-height: 1.1; }}
        .editor-picker-host {{ grid-column: 1 / -1; }}
        .editor-picker {{ display: none; gap: 6px; align-items: center; margin: 2px 0 0 0; justify-content: flex-start; flex-wrap: wrap; }}
        .editor-picker.open {{ display: flex; }}
        .editor-picker label {{ font-size: 0.8rem; color: #475569; }}
        .editor-picker input {{ border: 1px solid #cbd5e1; border-radius: 6px; padding: 3px 6px; font: inherit; }}
        .note-modal-backdrop {{
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.45);
            display: none;
            align-items: center;
            justify-content: center;
            z-index: 10020;
            padding: 16px;
        }}
        .note-modal-backdrop.open {{ display: flex; }}
        .note-modal {{
            width: min(760px, 96vw);
            background: #ffffff;
            border: 1px solid #cbd5e1;
            border-radius: 12px;
            box-shadow: 0 18px 44px rgba(15, 23, 42, 0.28);
            padding: 12px;
            display: grid;
            gap: 10px;
        }}
        .note-modal h3 {{
            margin: 0;
            font-size: 1rem;
            color: #0f172a;
        }}
        .note-modal textarea {{
            width: 100%;
            min-height: 180px;
            resize: vertical;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            padding: 8px 10px;
            font: inherit;
            color: #0f172a;
            background: #fff;
        }}
        .note-modal-grid {{
            display: grid;
            gap: 8px;
        }}
        .note-check-row {{
            display: flex;
            gap: 12px;
            flex-wrap: wrap;
        }}
        .note-check-row label {{
            display: inline-flex;
            gap: 6px;
            align-items: center;
            color: #334155;
            font-size: 0.9rem;
        }}
        .note-modal-grid input[type='text'] {{
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 7px 9px;
            font: inherit;
            color: #0f172a;
            background: #fff;
        }}
        .note-modal-actions {{
            display: flex;
            justify-content: flex-end;
            gap: 8px;
            flex-wrap: wrap;
        }}
        .editor-empty {{ color: #64748b; font-size: 0.9rem; }}
        .add-server {{
            margin-top: 12px;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #f8fafc;
            padding: 10px;
            display: grid;
            gap: 8px;
        }}
        .add-server-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
            gap: 8px;
            align-items: end;
        }}
        .add-server-grid label {{
            display: grid;
            gap: 4px;
            font-size: 0.82rem;
            color: #475569;
        }}
        .add-server-grid input,
        .add-server-grid select {{
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            padding: 5px 7px;
            font: inherit;
            color: #0f172a;
            background: #fff;
        }}
        .add-server-actions {{ display: flex; gap: 8px; align-items: center; flex-wrap: wrap; }}
        .add-server-divider {{
            height: 1px;
            background: #e2e8f0;
            margin: 4px 0;
        }}
        .fallback {{
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #ffffff;
            padding: 10px;
            overflow-x: auto;
        }}
        .fallback-grid {{ min-width: 760px; }}
        .fallback-row {{
            display: grid;
            grid-template-columns: minmax(240px, 1.2fr) minmax(480px, 3fr);
            gap: 10px;
            align-items: center;
            padding: 4px 0;
        }}
        .fallback-name {{
            font-size: 0.9rem;
            color: #0f172a;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .fallback-name-wrap {{
            display: flex;
            align-items: center;
            gap: 8px;
            min-width: 0;
        }}
        .fallback-track {{
            position: relative;
            height: 20px;
            border: 1px dashed #cbd5e1;
            border-radius: 8px;
            background: #f8fafc;
        }}
        .fallback-bar {{
            position: absolute;
            top: 3px;
            height: 12px;
            border-radius: 6px;
            background: #0ea5e9;
        }}
        .fallback-dates {{
            margin-top: 8px;
            color: #64748b;
            font-size: 0.82rem;
        }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
                        <div class="page-head">
                                <h1>{title_text}</h1>
                            <div class="head-actions">
{switch_link}
                                <button id="saveChangesBtn" class="btn-secondary" type="button">Save changes</button>
                                <button id="undoBtn" class="btn-secondary" type="button">Undo</button>
                                <button id="redoBtn" class="btn-secondary" type="button">Redo</button>
                                <button id="exportExcelBtnTop" class="btn-secondary" type="button">Export Excel</button>
                                <button id="exportConfluenceBtnTop" class="btn-secondary" type="button">Export Confluence</button>
                                <button id="exportConfluenceImageBtnTop" class="btn-secondary" type="button">Export Confluence image</button>
                                <button id="toggleFullscreen" class="btn-secondary btn-fullscreen" type="button">Maximize</button>
                            </div>
                        </div>
      <div class="meta-row">
          <div class="meta">Generated from Tidplan.xlsx</div>
          <div class="master-tools">
              <span id="masterBackupLabel" class="master-label">Master copy: {master_backup_label}</span>
              <a id="openMasterCopyLink" class="btn-secondary nav-link" href="backups/master_snapshot/{master_copy_file}" target="_blank" rel="noreferrer">Open master copy</a>
              <button id="restoreFromMasterBtn" class="btn-secondary" type="button">Load master copy</button>
              <button id="saveToMasterBtn" class="btn-secondary" type="button">Save to master copy</button>
          </div>
      </div>
            <section class="overview" aria-label="Overview">
                <div id="overviewCardTotal" class="overview-card" role="button" tabindex="0" aria-label="Show servers in total">
                    <span class="overview-label">Total</span>
                    <span id="overviewTotal" class="overview-value">0</span>
                </div>
                <div id="overviewCardDone" class="overview-card done" role="button" tabindex="0" aria-label="Show completed servers">
                    <span class="overview-label">Done</span>
                    <span id="overviewDone" class="overview-value">0</span>
                </div>
                <div id="overviewCardPartial" class="overview-card partial" role="button" tabindex="0" aria-label="Show partially completed servers">
                    <span class="overview-label">Partial</span>
                    <span id="overviewPartial" class="overview-value">0</span>
                </div>
                <div id="overviewCardRemaining" class="overview-card remaining" role="button" tabindex="0" aria-label="Show remaining servers">
                    <span class="overview-label">Remaining</span>
                    <span id="overviewRemaining" class="overview-value">0</span>
                </div>
            </section>
            <div class="controls">
{chr(10).join(button_items)}
      </div>
            <div class="filters">
                <div class="field">
                    <label for="fromDate">From</label>
                    <input id="fromDate" type="date" value="{frame_start_iso}" />
                </div>
                <div class="field">
                    <label for="toDate">To</label>
                    <input id="toDate" type="date" value="{last_date_iso}" />
                </div>
                <button id="resetFilters" class="btn-secondary" type="button">Reset</button>
            </div>
{planner_section}
                        <section class="add-server">
                            <strong>Add server to Gantt</strong>
                            <div class="add-server-grid">
                                <label>Country
                                    <select id="addCountry"></select>
                                </label>
                                <label>New country (optional)
                                    <input id="addCountryFree" type="text" placeholder="e.g. Denmark" />
                                </label>
                                <label>Server name
                                    <input id="addServerName" type="text" placeholder="e.g. se100new01" />
                                </label>
                                <label>Environment
                                    <input id="addEnv" type="text" placeholder="PROD/TEST" />
                                </label>
                                <label>Note (optional)
                                    <input id="addNote" type="text" placeholder="Short note" />
                                </label>
                                <label>From
                                    <input id="addFrom" type="date" value="{frame_start_iso}" />
                                </label>
                                <label>Days
                                    <input id="addDuration" type="number" min="1" max="60" step="1" value="1" />
                                </label>
                            </div>
                            <div class="add-server-actions">
                                <button id="addServerBtn" class="btn-secondary" type="button">Add server</button>
                                <button id="exportExcelBtn" class="btn-secondary" type="button">Download updated Excel</button>
                                <button id="exportConfluenceBtn" class="btn-secondary" type="button">Download Confluence export</button>
                                <span id="addServerMsg" class="status"></span>
                            </div>
                            <div class="add-server-divider"></div>
                            <strong>Remove server</strong>
                            <div class="add-server-grid">
                                <label>Country
                                    <select id="removeCountry"></select>
                                </label>
                                <label>Server
                                    <select id="removeServer"></select>
                                </label>
                            </div>
                            <div class="add-server-actions">
                                <button id="removeServerBtn" class="btn-secondary" type="button">Remove selected server</button>
                                <span id="removeServerMsg" class="status"></span>
                            </div>
                        </section>
                        <div id="statusText" class="status"></div>
                        <div id="chart"></div>
                                                <h2 class="editor-title">Drag editor (drag name or bar)</h2>
                                                <div class="editor-filters">
                                                    <div class="field">
                                                        <label for="dragFilterCountry">Country</label>
                                                        <select id="dragFilterCountry">
                                                            <option value="all">All</option>
                                                        </select>
                                                    </div>
                                                    <div class="field">
                                                        <label for="dragFilterServer">Server name</label>
                                                        <select id="dragFilterServer">
                                                            <option value="all">All</option>
                                                        </select>
                                                    </div>
                                                    <div class="field">
                                                        <label for="dragFilterEnvironment">Environment</label>
                                                        <select id="dragFilterEnvironment">
                                                            <option value="all">All</option>
                                                        </select>
                                                    </div>
                                                    <div class="field">
                                                        <label for="dragFilterFrom">From</label>
                                                        <input id="dragFilterFrom" type="date" />
                                                    </div>
                                                    <div class="field">
                                                        <label for="dragFilterTo">To</label>
                                                        <input id="dragFilterTo" type="date" />
                                                    </div>
                                                    <button id="dragFilterReset" class="btn-secondary" type="button">Reset drag filters</button>
                                                </div>
                                                <div id="dragEditor" class="editor"></div>
                        <div id="noteModalBackdrop" class="note-modal-backdrop" aria-hidden="true">
                            <div class="note-modal" role="dialog" aria-modal="true" aria-labelledby="noteModalTitle">
                                <h3 id="noteModalTitle">Note</h3>
                                <div class="note-modal-grid">
                                    <div class="note-check-row">
                                        <label><input id="noteRhelOk" type="checkbox" /> RHEL9 ok</label>
                                        <label><input id="noteImeOk" type="checkbox" /> IME V11 ok</label>
                                        <label><input id="noteAppTestOk" type="checkbox" /> Application test ok</label>
                                        <label><input id="noteRegressionTestOk" type="checkbox" /> Regression test ok</label>
                                        <label><input id="noteIpSwapOk" type="checkbox" /> IP Swap ok</label>
                                        <label><input id="notePatchOk" type="checkbox" /> Patching ok</label>
                                    </div>
                                    <input id="noteIpSwapStartDate" type="date" placeholder="IP Swap start date" />
                                    <input id="noteIpSwapEndDate" type="date" placeholder="IP Swap end date" />
                                    <input id="noteIpSwapFrom" type="text" placeholder="Swap from server" />
                                    <input id="noteIpSwapTo" type="text" placeholder="Swap to server" />
                                    <input id="notePatchNumber" type="text" placeholder="Patch number" />
                                </div>
                                <textarea id="noteModalInput" placeholder="Write a note for the server..."></textarea>
                                <div class="note-modal-actions">
                                    <button id="noteModalClear" class="btn-secondary" type="button">Clear</button>
                                    <button id="noteModalCancel" class="btn-secondary" type="button">Close</button>
                                    <button id="noteModalSave" class="btn-secondary" type="button">Save</button>
                                </div>
                            </div>
                        </div>
                        <div id="overviewModalBackdrop" class="overview-modal-backdrop" aria-hidden="true">
                            <div class="overview-modal" role="dialog" aria-modal="true" aria-labelledby="overviewModalTitle">
                                <h3 id="overviewModalTitle">Servers</h3>
                                <p id="overviewModalEmpty" class="overview-empty" hidden>No servers in this category.</p>
                                <ol id="overviewModalList" class="overview-modal-list"></ol>
                                <div class="overview-modal-actions">
                                    <button id="overviewModalClose" class="btn-secondary" type="button">Close</button>
                                </div>
                            </div>
                        </div>
    </div>
  </div>
    <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.min.js"></script>
    <script>
        const mermaid = window.mermaid;

        const countries = {countries_json};
        const rows = {rows_json};
        const defaultFrom = '{frame_start_iso}';
        const defaultTo = '{last_date_iso}';

        const buttons = Array.from(document.querySelectorAll('.country-btn'));
        const fromInput = document.getElementById('fromDate');
        const toInput = document.getElementById('toDate');
        const resetBtn = document.getElementById('resetFilters');
        const planStartDate = document.getElementById('planStartDate');
        const planPerDay = document.getElementById('planPerDay');
        const planSkipWeekends = document.getElementById('planSkipWeekends');
        const planSuggestBtn = document.getElementById('planSuggestBtn');
        const chart = document.getElementById('chart');
        const statusText = document.getElementById('statusText');
        const dragEditor = document.getElementById('dragEditor');
        const dragFilterCountry = document.getElementById('dragFilterCountry');
        const dragFilterServer = document.getElementById('dragFilterServer');
        const dragFilterEnvironment = document.getElementById('dragFilterEnvironment');
        const dragFilterFrom = document.getElementById('dragFilterFrom');
        const dragFilterTo = document.getElementById('dragFilterTo');
        const dragFilterReset = document.getElementById('dragFilterReset');
        const noteModalBackdrop = document.getElementById('noteModalBackdrop');
        const noteModalTitle = document.getElementById('noteModalTitle');
        const noteRhelOk = document.getElementById('noteRhelOk');
        const noteImeOk = document.getElementById('noteImeOk');
        const noteAppTestOk = document.getElementById('noteAppTestOk');
        const noteRegressionTestOk = document.getElementById('noteRegressionTestOk');
        const noteIpSwapOk = document.getElementById('noteIpSwapOk');
        const noteIpSwapStartDate = document.getElementById('noteIpSwapStartDate');
        const noteIpSwapEndDate = document.getElementById('noteIpSwapEndDate');
        const noteIpSwapFrom = document.getElementById('noteIpSwapFrom');
        const noteIpSwapTo = document.getElementById('noteIpSwapTo');
        const notePatchOk = document.getElementById('notePatchOk');
        const notePatchNumber = document.getElementById('notePatchNumber');
        const noteModalInput = document.getElementById('noteModalInput');
        const noteModalSave = document.getElementById('noteModalSave');
        const noteModalCancel = document.getElementById('noteModalCancel');
        const noteModalClear = document.getElementById('noteModalClear');
        const addCountry = document.getElementById('addCountry');
        const addCountryFree = document.getElementById('addCountryFree');
        const addServerName = document.getElementById('addServerName');
        const addEnv = document.getElementById('addEnv');
        const addNote = document.getElementById('addNote');
        const addFrom = document.getElementById('addFrom');
        const addDuration = document.getElementById('addDuration');
        const addServerBtn = document.getElementById('addServerBtn');
        const exportExcelBtn = document.getElementById('exportExcelBtn');
        const exportExcelBtnTop = document.getElementById('exportExcelBtnTop');
        const exportConfluenceBtn = document.getElementById('exportConfluenceBtn');
        const exportConfluenceBtnTop = document.getElementById('exportConfluenceBtnTop');
        const exportConfluenceImageBtnTop = document.getElementById('exportConfluenceImageBtnTop');
        const addServerMsg = document.getElementById('addServerMsg');
        const toggleFullscreenBtn = document.getElementById('toggleFullscreen');
        const saveChangesBtn = document.getElementById('saveChangesBtn');
        const undoBtn = document.getElementById('undoBtn');
        const redoBtn = document.getElementById('redoBtn');
        const removeCountry = document.getElementById('removeCountry');
        const removeServer = document.getElementById('removeServer');
        const removeServerBtn = document.getElementById('removeServerBtn');
        const removeServerMsg = document.getElementById('removeServerMsg');
        const masterBackupLabel = document.getElementById('masterBackupLabel');
        const openMasterCopyLink = document.getElementById('openMasterCopyLink');
        const restoreFromMasterBtn = document.getElementById('restoreFromMasterBtn');
        const saveToMasterBtn = document.getElementById('saveToMasterBtn');
        const overviewCardTotal = document.getElementById('overviewCardTotal');
        const overviewCardDone = document.getElementById('overviewCardDone');
        const overviewCardPartial = document.getElementById('overviewCardPartial');
        const overviewCardRemaining = document.getElementById('overviewCardRemaining');
        const overviewTotal = document.getElementById('overviewTotal');
        const overviewDone = document.getElementById('overviewDone');
        const overviewPartial = document.getElementById('overviewPartial');
        const overviewRemaining = document.getElementById('overviewRemaining');
        const overviewModalBackdrop = document.getElementById('overviewModalBackdrop');
        const overviewModalTitle = document.getElementById('overviewModalTitle');
        const overviewModalEmpty = document.getElementById('overviewModalEmpty');
        const overviewModalList = document.getElementById('overviewModalList');
        const overviewModalClose = document.getElementById('overviewModalClose');

        const STORAGE_KEY = 'gantt_plan_saved_state_v1';
        const HISTORY_LIMIT = 80;
        const undoStack = [];
        const redoStack = [];

        let selectedCountry = 'all';
        let removeCandidates = [];
        let activeNoteRow = null;
        let activeNoteBadgeUpdater = null;
        const locationProbe = [window.location.href || '', window.location.pathname || '', document.baseURI || '']
            .join(' ')
            .split('\\\\')
            .join('/')
            .toLowerCase();
        const isPlanningView = locationProbe.includes('tidplan_gantt_kalkyl.html');
        const currentViewFileName = isPlanningView ? 'tidplan_gantt_kalkyl.html' : 'tidplan_gantt_status_per_land.html';
        const isMasterSnapshotView = locationProbe.includes('master_snapshot');

        if (isMasterSnapshotView) {{
            [openMasterCopyLink, restoreFromMasterBtn, saveToMasterBtn].forEach((node) => {{
                if (!node) {{
                    return;
                }}
                node.style.display = 'none';
                node.setAttribute('aria-hidden', 'true');
            }});
        }}

        mermaid.initialize({{ startOnLoad: false, theme: 'default', securityLevel: 'loose' }});

        function slug(value) {{
            return value.toLowerCase().replace(/[^a-z0-9]+/g, '_').replace(/^_|_$/g, '');
        }}

        function updateActiveButton() {{
            buttons.forEach((btn) => btn.classList.toggle('active', btn.dataset.country === selectedCountry));
        }}

        function createCountryButton(country) {{
            const btn = document.createElement('button');
            btn.className = 'country-btn';
            btn.type = 'button';
            btn.dataset.country = country;
            btn.textContent = country;
            btn.addEventListener('click', () => {{
                selectedCountry = country;
                updateActiveButton();
                applyCountryDateRange(country);
                refreshCountrySelect();
                renderChart();
            }});
            const controls = document.querySelector('.controls');
            controls?.appendChild(btn);
            buttons.push(btn);
            updateActiveButton();
        }}

        function removeCountryButton(country) {{
            const index = buttons.findIndex((btn) => btn.dataset.country === country);
            if (index < 0) {{
                return;
            }}
            const btn = buttons[index];
            buttons.splice(index, 1);
            btn.remove();
        }}

        function syncCountryButtons() {{
            const existing = buttons
                .map((btn) => btn.dataset.country)
                .filter((country) => country && country !== 'all');

            existing.forEach((country) => {{
                if (!countries.includes(country)) {{
                    removeCountryButton(country);
                }}
            }});

            countries.forEach((country) => {{
                if (!buttons.some((btn) => btn.dataset.country === country)) {{
                    createCountryButton(country);
                }}
            }});

            refreshDragFilterOptions();
        }}

        function refreshDragFilterOptions() {{
            if (!dragFilterCountry || !dragFilterServer || !dragFilterEnvironment || !dragFilterFrom || !dragFilterTo) {{
                return;
            }}

            const previousCountry = String(dragFilterCountry.value || 'all');
            const previousServer = String(dragFilterServer.value || 'all');
            const previousEnvironment = String(dragFilterEnvironment.value || 'all');
            const previousFrom = String(dragFilterFrom.value || '');
            const previousTo = String(dragFilterTo.value || '');

            dragFilterCountry.innerHTML = '';
            dragFilterServer.innerHTML = '';
            dragFilterEnvironment.innerHTML = '';

            const allOption = document.createElement('option');
            allOption.value = 'all';
            allOption.textContent = 'All';
            dragFilterCountry.appendChild(allOption);

            const allServerOption = document.createElement('option');
            allServerOption.value = 'all';
            allServerOption.textContent = 'All';
            dragFilterServer.appendChild(allServerOption);

            const allEnvironmentOption = document.createElement('option');
            allEnvironmentOption.value = 'all';
            allEnvironmentOption.textContent = 'All';
            dragFilterEnvironment.appendChild(allEnvironmentOption);

            countries.forEach((country) => {{
                const option = document.createElement('option');
                option.value = country;
                option.textContent = country;
                dragFilterCountry.appendChild(option);
            }});

            const servers = Array.from(new Set(rows.map((r) => String(r.server || '').trim()).filter((v) => v))).sort((a, b) => a.localeCompare(b));
            servers.forEach((serverName) => {{
                const option = document.createElement('option');
                option.value = serverName;
                option.textContent = serverName;
                dragFilterServer.appendChild(option);
            }});

            const environments = Array.from(new Set(rows.map((r) => String(r.environment || '').trim()).filter((v) => v))).sort((a, b) => a.localeCompare(b));
            environments.forEach((environment) => {{
                const option = document.createElement('option');
                option.value = environment;
                option.textContent = environment;
                dragFilterEnvironment.appendChild(option);
            }});

            dragFilterCountry.value = previousCountry === 'all' || countries.includes(previousCountry) ? previousCountry : 'all';
            dragFilterServer.value = previousServer === 'all' || servers.includes(previousServer) ? previousServer : 'all';
            dragFilterEnvironment.value = previousEnvironment === 'all' || environments.includes(previousEnvironment) ? previousEnvironment : 'all';
            dragFilterFrom.value = previousFrom;
            dragFilterTo.value = previousTo;
        }}

        function cloneRows(sourceRows) {{
            return sourceRows.map((r) => ({{
                priority: Number.isFinite(Number(r.priority)) ? Number(r.priority) : 9999,
                country: String(r.country || ''),
                server: String(r.server || ''),
                environment: String(r.environment || ''),
                date: String(r.date || defaultFrom),
                duration: Math.max(1, Number(r.duration || 1)),
                status: normalizeProgressStatus(String(r.status || 'none')),
                rhelOk: Boolean(r.rhelOk),
                imeOk: Boolean(r.imeOk),
                appTestOk: Boolean(r.appTestOk),
                regressionTestOk: Boolean(r.regressionTestOk),
                ipSwapOk: Boolean(r.ipSwapOk),
                ipSwapStartDate: String(r.ipSwapStartDate || ''),
                ipSwapEndDate: String(r.ipSwapEndDate || r.ipSwapStartDate || ''),
                ipSwapFrom: String(r.ipSwapFrom || ''),
                ipSwapTo: String(r.ipSwapTo || ''),
                patchOk: Boolean(r.patchOk),
                patchNumber: String(r.patchNumber || ''),
                notesProd: String(r.notesProd || r.notes || ''),
                notesTest: String(r.notesTest || ''),
                notes: String(r.notes || ''),
            }}));
        }}

        function normalizeProgressStatus(value) {{
            const raw = String(value || '').toLowerCase().trim();
            if (raw === 'done' || raw === 'green' || raw === 'klar' || raw === 'complete' || raw === 'completed' || raw === 'full') {{
                return 'done';
            }}
            if (raw === 'partial' || raw === 'half' || raw === 'halv' || raw === 'delvis') {{
                return 'partial';
            }}
            return 'none';
        }}

        function nextProgressStatus(current) {{
            if (current === 'none') {{
                return 'partial';
            }}
            if (current === 'partial') {{
                return 'done';
            }}
            return 'none';
        }}

        function progressStatusLabel(current) {{
            if (current === 'done') {{
                return 'Done (green)';
            }}
            if (current === 'partial') {{
                return 'Partial (half green/half red)';
            }}
            return 'Not done (red)';
        }}

        const lifecycleSegments = [
            {{ key: 'rhelOk', label: 'RHEL9', color: '#22c55e' }},
            {{ key: 'imeOk', label: 'IME V11', color: '#22c55e' }},
            {{ key: 'appTestOk', label: 'Application test', color: '#22c55e' }},
            {{ key: 'regressionTestOk', label: 'Regression test', color: '#22c55e' }},
            {{ key: 'ipSwapOk', label: 'IP Swap', color: '#22c55e' }},
            {{ key: 'patchOk', label: 'Patching', color: '#22c55e' }},
        ];

        function getLifecycleInfo(rowRef) {{
            const done = [];
            const pending = [];
            lifecycleSegments.forEach((segment) => {{
                if (Boolean(rowRef[segment.key])) {{
                    done.push(segment.label);
                }} else {{
                    pending.push(segment.label);
                }}
            }});
            return {{ done, pending }};
        }}

        function lifecycleGradient(rowRef) {{
            const segmentCount = lifecycleSegments.length;
            const gradientStops = lifecycleSegments.map((segment, idx) => {{
                const from = (idx / segmentCount) * 100;
                const to = ((idx + 1) / segmentCount) * 100;
                const fill = Boolean(rowRef[segment.key]) ? segment.color : '#ef4444';
                return `${{fill}} ${{from}}%, ${{fill}} ${{to}}%`;
            }}).join(', ');
            return `linear-gradient(90deg, ${{gradientStops}})`;
        }}

        function lifecycleHoverText(rowRef) {{
            const info = getLifecycleInfo(rowRef);
            const fromIso = rowRef.date;
            const toIso = getEndDate(rowRef.date, rowRef.duration);
            const serverLabel = buildServerLabel(rowRef);
            const doneText = info.done.length ? `Upgraded: ${{info.done.join(', ')}}` : 'Upgraded: none';
            const pendingText = info.pending.length ? `Pending: ${{info.pending.join(', ')}}` : 'Pending: none';
            const ipSwapWindowText = rowRef.ipSwapStartDate || rowRef.ipSwapEndDate
                ? `IP Swap plan: ${{String(rowRef.ipSwapStartDate || rowRef.ipSwapEndDate || '-')}} -> ${{String(rowRef.ipSwapEndDate || rowRef.ipSwapStartDate || '-')}}`
                : '';
            const ipSwapRoute = rowRef.ipSwapFrom || rowRef.ipSwapTo
                ? `IP Swap: ${{String(rowRef.ipSwapFrom || '-')}} -> ${{String(rowRef.ipSwapTo || '-')}}`
                : '';
            const patchText = rowRef.patchNumber ? `Patch number: ${{rowRef.patchNumber}}` : '';
            const noteText = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '').trim();
            return [
                `Server: ${{serverLabel}}`,
                `${{fromIso}} -> ${{toIso}}`,
                doneText,
                pendingText,
                ipSwapWindowText,
                ipSwapRoute,
                patchText,
                noteText ? `Note: ${{noteText}}` : '',
            ].filter(Boolean).join('\\n');
        }}

        function snapshotState() {{
            return {{
                rows: cloneRows(rows),
                countries: [...countries],
                selectedCountry,
                fromDate: fromInput?.value || defaultFrom,
                toDate: toInput?.value || defaultTo,
            }};
        }}

        function updateUndoButton() {{
            if (!undoBtn) {{
                return;
            }}
            const count = undoStack.length;
            undoBtn.disabled = count === 0;
            undoBtn.textContent = count > 0 ? `Undo (${{count}})` : 'Undo';
        }}

        function updateRedoButton() {{
            if (!redoBtn) {{
                return;
            }}
            const count = redoStack.length;
            redoBtn.disabled = count === 0;
            redoBtn.textContent = count > 0 ? `Redo (${{count}})` : 'Redo';
        }}

        function pushUndoSnapshot(clearRedo = true) {{
            undoStack.push(snapshotState());
            if (undoStack.length > HISTORY_LIMIT) {{
                undoStack.shift();
            }}
            if (clearRedo) {{
                redoStack.length = 0;
            }}
            updateUndoButton();
            updateRedoButton();
        }}

        function applyState(nextState) {{
            if (!nextState || !Array.isArray(nextState.rows) || !Array.isArray(nextState.countries)) {{
                return false;
            }}

            rows.length = 0;
            cloneRows(nextState.rows).forEach((r) => rows.push(r));

            countries.length = 0;
            nextState.countries
                .map((c) => String(c || '').trim())
                .filter((c) => c)
                .forEach((c) => countries.push(c));

            syncCountryButtons();

            const nextSelected = String(nextState.selectedCountry || 'all');
            selectedCountry = nextSelected === 'all' || countries.includes(nextSelected) ? nextSelected : 'all';

            if (fromInput) {{
                fromInput.value = nextState.fromDate || defaultFrom;
            }}
            if (toInput) {{
                toInput.value = nextState.toDate || defaultTo;
            }}
            if (fromInput?.value && toInput?.value && fromInput.value > toInput.value) {{
                toInput.value = fromInput.value;
            }}

            updateActiveButton();
            refreshCountrySelect();
            ensureDateBoundsCoverRows();
            renderChart();
            return true;
        }}

        function saveCurrentState() {{
            try {{
                const state = snapshotState();
                const payload = {{
                    version: 1,
                    savedAt: new Date().toISOString(),
                    state: state,
                }};
                const jsonStr = JSON.stringify(payload);
                window.localStorage.setItem(STORAGE_KEY, jsonStr);
                statusText.classList.remove('error');
                statusText.textContent = `Changes saved locally in the browser. (${{Math.round(jsonStr.length / 1024)}}KB used)`;
                console.log('State saved successfully. From:', state.fromDate, 'To:', state.toDate);
            }} catch (err) {{
                statusText.classList.add('error');
                const errorMsg = err.message ? ` (${{err.message}})` : '';
                statusText.textContent = `Could not save locally. Storage full or permissions issue.${{errorMsg}}`;
                console.error('Failed to save state:', err);
            }}
        }}

        function padTwo(value) {{
            return String(value).padStart(2, '0');
        }}

        function formatLocalStamp(dateValue) {{
            return `${{dateValue.getFullYear()}}-${{padTwo(dateValue.getMonth() + 1)}}-${{padTwo(dateValue.getDate())}} ${{padTwo(dateValue.getHours())}}:${{padTwo(dateValue.getMinutes())}}:${{padTwo(dateValue.getSeconds())}}`;
        }}

        function buildCurrentHtmlSnapshot() {{
            let html = document.documentElement.outerHTML;
            const rowsJson = JSON.stringify(rows);
            const countriesJson = JSON.stringify(countries);
            const currentFrom = String(fromInput?.value || defaultFrom);
            const currentTo = String(toInput?.value || defaultTo);

            html = html.replace(/const countries = .*?;/, `const countries = ${{countriesJson}};`);
            html = html.replace(/const rows = .*?;/, `const rows = ${{rowsJson}};`);
            html = html.replace(/const defaultFrom = '.*?';/, `const defaultFrom = '${{currentFrom}}';`);
            html = html.replace(/const defaultTo = '.*?';/, `const defaultTo = '${{currentTo}}';`);
            return `<!doctype html>\n${{html}}`;
        }}

        async function writeDirectoryTextFile(dirHandle, fileName, content) {{
            const fileHandle = await dirHandle.getFileHandle(fileName, {{ create: true }});
            const writable = await fileHandle.createWritable();
            await writable.write(content);
            await writable.close();
        }}

        function downloadTextFile(fileName, content, mimeType = 'text/plain;charset=utf-8') {{
            const blob = new Blob([content], {{ type: mimeType }});
            const url = URL.createObjectURL(blob);
            const link = document.createElement('a');
            link.href = url;
            link.download = fileName;
            link.style.display = 'none';
            document.body.appendChild(link);
            link.click();
            link.remove();
            setTimeout(() => URL.revokeObjectURL(url), 1500);
        }}

        function saveToMasterCopyByDownload(autoDownload = true) {{
            const savedAt = formatLocalStamp(new Date());
            const markerText = `master_snapshot refreshed: ${{savedAt}}\nPurpose: stable master baseline separate from latest_snapshot\n`;
            if (autoDownload) {{
                downloadTextFile(currentViewFileName, buildCurrentHtmlSnapshot(), 'text/html;charset=utf-8');
                downloadTextFile('MASTER_BACKUP.txt', markerText);
            }}
            if (masterBackupLabel) {{
                masterBackupLabel.textContent = `Master copy: master_snapshot refreshed: ${{savedAt}}`;
            }}
            statusText.classList.remove('error');
            statusText.textContent = autoDownload
                ? `Master copy files downloaded (${{currentViewFileName}} + MASTER_BACKUP.txt). Move them to backups/master_snapshot.`
                : 'Could not save master copy directly in this view.';
        }}

        async function resolveMasterSnapshotDir(baseDirHandle) {{
            if (!baseDirHandle) {{
                throw new Error('No directory handle selected');
            }}

            if (baseDirHandle.name === 'master_snapshot') {{
                return {{ masterDir: baseDirHandle, backupsDir: null, pathHint: 'master_snapshot' }};
            }}

            if (baseDirHandle.name === 'backups') {{
                const masterDir = await baseDirHandle.getDirectoryHandle('master_snapshot', {{ create: true }});
                return {{ masterDir, backupsDir: baseDirHandle, pathHint: 'backups/master_snapshot' }};
            }}

            try {{
                const backupsDir = await baseDirHandle.getDirectoryHandle('backups', {{ create: true }});
                const masterDir = await backupsDir.getDirectoryHandle('master_snapshot', {{ create: true }});
                return {{ masterDir, backupsDir, pathHint: 'backups/master_snapshot' }};
            }} catch (_err) {{
                // Fallback when selected folder is not repo root or nested folder creation is blocked.
                return {{ masterDir: baseDirHandle, backupsDir: null, pathHint: `selected folder (${{baseDirHandle.name || 'unknown'}})` }};
            }}
        }}

        async function saveToMasterCopy() {{
            if (!window.showDirectoryPicker) {{
                saveToMasterCopyByDownload(true);
                return;
            }}

            try {{
                const dirHandle = await window.showDirectoryPicker({{ mode: 'readwrite' }});
                const resolved = await resolveMasterSnapshotDir(dirHandle);
                const savedAt = formatLocalStamp(new Date());
                await writeDirectoryTextFile(resolved.masterDir, currentViewFileName, buildCurrentHtmlSnapshot());
                await writeDirectoryTextFile(
                    resolved.masterDir,
                    'MASTER_BACKUP.txt',
                    `master_snapshot refreshed: ${{savedAt}}\nPurpose: stable master baseline separate from latest_snapshot\n`
                );
                if (resolved.backupsDir) {{
                    try {{
                        await writeDirectoryTextFile(
                            resolved.backupsDir,
                            'MASTER_BACKUP.txt',
                            `master_snapshot refreshed: ${{savedAt}}\nPurpose: stable master baseline separate from latest_snapshot\n`
                        );
                    }} catch (_markerErr) {{
                        // Non-critical marker write should not block successful master snapshot save.
                    }}
                }}
                if (masterBackupLabel) {{
                    masterBackupLabel.textContent = `Master copy: master_snapshot refreshed: ${{savedAt}}`;
                }}
                statusText.classList.remove('error');
                statusText.textContent = `Current view saved to ${{resolved.pathHint}}.`;
            }} catch (err) {{
                if (err && err.name === 'AbortError') {{
                    statusText.classList.remove('error');
                    statusText.textContent = 'Save to master copy was cancelled.';
                    return;
                }}
                if (err && err.name === 'SecurityError') {{
                    saveToMasterCopyByDownload(true);
                    return;
                }}
                statusText.classList.add('error');
                const errName = err && err.name ? String(err.name) : 'UnknownError';
                const errMsg = err && err.message ? String(err.message) : 'Unknown error';
                statusText.textContent = `Could not save to master copy (${{errName}}: ${{errMsg}}).`;
            }}
        }}

        function extractJsonArrayAssignment(sourceText, variableName) {{
            const marker = `const ${{variableName}} = `;
            const markerIndex = sourceText.indexOf(marker);
            if (markerIndex < 0) {{
                return null;
            }}

            const startIndex = sourceText.indexOf('[', markerIndex + marker.length);
            if (startIndex < 0) {{
                return null;
            }}

            let depth = 0;
            let inString = false;
            let quoteChar = '';
            let escaped = false;

            for (let i = startIndex; i < sourceText.length; i += 1) {{
                const ch = sourceText[i];
                if (inString) {{
                    if (escaped) {{
                        escaped = false;
                        continue;
                    }}
                    if (ch === '\\\\\\\\') {{
                        escaped = true;
                        continue;
                    }}
                    if (ch === quoteChar) {{
                        inString = false;
                    }}
                    continue;
                }}

                if (ch === '"' || ch === "'") {{
                    inString = true;
                    quoteChar = ch;
                    continue;
                }}

                if (ch === '[') {{
                    depth += 1;
                    continue;
                }}
                if (ch === ']') {{
                    depth -= 1;
                    if (depth === 0) {{
                        return sourceText.slice(startIndex, i + 1);
                    }}
                }}
            }}

            return null;
        }}

        function extractStringAssignment(sourceText, variableName) {{
            const marker = `const ${{variableName}} = '`;
            const start = sourceText.indexOf(marker);
            if (start < 0) {{
                return '';
            }}
            let i = start + marker.length;
            let value = '';
            let escaped = false;
            for (; i < sourceText.length; i += 1) {{
                const ch = sourceText[i];
                if (escaped) {{
                    value += ch;
                    escaped = false;
                    continue;
                }}
                if (ch === '\\\\\\\\') {{
                    escaped = true;
                    continue;
                }}
                if (ch === "'") {{
                    break;
                }}
                value += ch;
            }}
            return value;
        }}

        function parseMasterSnapshotState(htmlText) {{
            const countriesJson = extractJsonArrayAssignment(htmlText, 'countries');
            const rowsJson = extractJsonArrayAssignment(htmlText, 'rows');
            if (!countriesJson || !rowsJson) {{
                throw new Error('missing-state');
            }}

            const parsedCountries = JSON.parse(countriesJson);
            const parsedRows = JSON.parse(rowsJson);
            const parsedFrom = extractStringAssignment(htmlText, 'defaultFrom') || defaultFrom;
            const parsedTo = extractStringAssignment(htmlText, 'defaultTo') || defaultTo;

            return {{
                rows: Array.isArray(parsedRows) ? parsedRows : [],
                countries: Array.isArray(parsedCountries) ? parsedCountries : [],
                selectedCountry: 'all',
                fromDate: parsedFrom,
                toDate: parsedTo,
            }};
        }}

        async function tryReadMasterSnapshotText() {{
            try {{
                const response = await fetch(`backups/master_snapshot/${{currentViewFileName}}`, {{ cache: 'no-store' }});
                if (response && response.ok) {{
                    return await response.text();
                }}
            }} catch (_ignored) {{
            }}

            if (!window.showOpenFilePicker) {{
                throw new Error('no-open-file-picker');
            }}

            const handles = await window.showOpenFilePicker({{
                multiple: false,
                types: [{{
                    description: 'Master copy HTML',
                    accept: {{ 'text/html': ['.html', '.htm'] }},
                }}],
            }});
            if (!handles || !handles.length) {{
                throw new Error('no-file-selected');
            }}

            const file = await handles[0].getFile();
            return await file.text();
        }}

        async function restoreFromMasterCopy() {{
            try {{
                const masterHtml = await tryReadMasterSnapshotText();
                const importedState = parseMasterSnapshotState(masterHtml);
                pushUndoSnapshot();
                const applied = applyState(importedState);
                if (!applied) {{
                    throw new Error('apply-failed');
                }}
                statusText.classList.remove('error');
                statusText.textContent = 'Master copy loaded into the editor view.';
            }} catch (err) {{
                if (err && err.name === 'AbortError') {{
                    statusText.classList.remove('error');
                    statusText.textContent = 'Loading master copy was cancelled.';
                    return;
                }}
                statusText.classList.add('error');
                statusText.textContent = 'Could not load the master copy into the editor view.';
            }}
        }}

        function loadSavedState() {{
            try {{
                const raw = window.localStorage.getItem(STORAGE_KEY);
                if (!raw) {{
                    return false;
                }}
                const payload = JSON.parse(raw);
                if (!payload || !payload.state) {{
                    return false;
                }}
                return applyState(payload.state);
            }} catch (err) {{
                return false;
            }}
        }}

        function undoLastChange() {{
            if (!undoStack.length) {{
                statusText.classList.remove('error');
                statusText.textContent = 'No more steps to undo.';
                updateUndoButton();
                updateRedoButton();
                return;
            }}

            const previous = undoStack.pop();
            redoStack.push(snapshotState());
            if (redoStack.length > HISTORY_LIMIT) {{
                redoStack.shift();
            }}
            const restored = applyState(previous);
            updateUndoButton();
            updateRedoButton();
            if (restored) {{
                statusText.classList.remove('error');
                statusText.textContent = 'Undo complete.';
            }} else {{
                statusText.classList.add('error');
                statusText.textContent = 'Undo failed.';
            }}
        }}

        function redoLastChange() {{
            if (!redoStack.length) {{
                statusText.classList.remove('error');
                statusText.textContent = 'No more steps to redo.';
                updateUndoButton();
                updateRedoButton();
                return;
            }}

            const next = redoStack.pop();
            pushUndoSnapshot(false);
            const restored = applyState(next);
            updateUndoButton();
            updateRedoButton();
            if (restored) {{
                statusText.classList.remove('error');
                statusText.textContent = 'Redo complete.';
            }} else {{
                statusText.classList.add('error');
                statusText.textContent = 'Redo failed.';
            }}
        }}

        function getServerLabel(rowRef) {{
            return rowRef.environment ? `${{rowRef.server}} (${{rowRef.environment}})` : rowRef.server;
        }}

        function refreshRemoveServerSelect() {{
            if (!removeCountry || !removeServer) {{
                return;
            }}

            const chosenCountry = removeCountry.value;
            removeCandidates = rows.filter((r) => r.country === chosenCountry);
            removeServer.innerHTML = '';

            if (!removeCandidates.length) {{
                const opt = document.createElement('option');
                opt.value = '';
                opt.textContent = 'No servers';
                removeServer.appendChild(opt);
                return;
            }}

            removeCandidates.forEach((rowRef, idx) => {{
                const opt = document.createElement('option');
                opt.value = String(idx);
                opt.textContent = `${{getServerLabel(rowRef)}} | ${{rowRef.date}}`;
                removeServer.appendChild(opt);
            }});
        }}

        function refreshCountrySelect() {{
            if (!addCountry) {{
                return;
            }}
            const current = addCountry.value;
            addCountry.innerHTML = '';
            countries.forEach((country) => {{
                const opt = document.createElement('option');
                opt.value = country;
                opt.textContent = country;
                addCountry.appendChild(opt);
            }});

            if (removeCountry) {{
                const removeCurrent = removeCountry.value;
                removeCountry.innerHTML = '';
                countries.forEach((country) => {{
                    const opt = document.createElement('option');
                    opt.value = country;
                    opt.textContent = country;
                    removeCountry.appendChild(opt);
                }});
                if (selectedCountry !== 'all' && countries.includes(selectedCountry)) {{
                    removeCountry.value = selectedCountry;
                }} else if (countries.includes(removeCurrent)) {{
                    removeCountry.value = removeCurrent;
                }}
            }}

            if (selectedCountry !== 'all' && countries.includes(selectedCountry)) {{
                addCountry.value = selectedCountry;
            }} else if (countries.includes(current)) {{
                addCountry.value = current;
            }}

            refreshRemoveServerSelect();
        }}

        function normalizeLabel(value) {{
            return value.replace(/\s+/g, ' ').trim();
        }}

        function sanitizeMermaidLabel(value) {{
            return normalizeLabel(String(value || '')).replace(/[:,;]/g, '-');
        }}

        function countryShort(country) {{
            const key = (country || '').trim().toLowerCase();
            const mapping = {{
                sweden: 'swe',
                norway: 'nor',
                finland: 'fin',
                lithuania: 'ltu',
            }};
            if (mapping[key]) {{
                return mapping[key];
            }}
            const fallback = key.replace(/[^a-z]/g, '').slice(0, 3);
            return (fallback || 'n/a').toLowerCase();
        }}

        function buildServerLabel(rowRef) {{
            const base = rowRef.environment ? `${{rowRef.server}} (${{rowRef.environment}})` : rowRef.server;
            return `${{base}} (${{countryShort(rowRef.country)}})`;
        }}

        function parseIsoDate(value) {{
            const [year, month, day] = value.split('-').map((v) => Number(v));
            return new Date(Date.UTC(year, month - 1, day));
        }}

        function formatIsoDate(dateValue) {{
            const y = dateValue.getUTCFullYear();
            const m = String(dateValue.getUTCMonth() + 1).padStart(2, '0');
            const d = String(dateValue.getUTCDate()).padStart(2, '0');
            return `${{y}}-${{m}}-${{d}}`;
        }}

        function addDays(isoDate, days) {{
            const dateValue = parseIsoDate(isoDate);
            dateValue.setUTCDate(dateValue.getUTCDate() + days);
            return formatIsoDate(dateValue);
        }}

        function isWeekend(isoDate) {{
            const d = parseIsoDate(isoDate);
            const day = d.getUTCDay();
            return day === 0 || day === 6;
        }}

        function nextAllowedDay(isoDate, skipWeekends) {{
            let current = isoDate;
            while (skipWeekends && isWeekend(current)) {{
                current = addDays(current, 1);
            }}
            return current;
        }}

        function addWorkingDays(startIso, days, skipWeekends) {{
            let current = nextAllowedDay(startIso, skipWeekends);
            let remaining = Math.max(1, Number(days || 1)) - 1;
            while (remaining > 0) {{
                current = addDays(current, 1);
                current = nextAllowedDay(current, skipWeekends);
                remaining -= 1;
            }}
            return current;
        }}

        function clampDate(isoDate, minDate, maxDate) {{
            if (minDate && isoDate < minDate) {{
                return minDate;
            }}
            if (maxDate && isoDate > maxDate) {{
                return maxDate;
            }}
            return isoDate;
        }}

        function daysBetween(startIso, endIso) {{
            const msPerDay = 24 * 60 * 60 * 1000;
            const a = parseIsoDate(startIso);
            const b = parseIsoDate(endIso);
            return Math.round((b.getTime() - a.getTime()) / msPerDay);
        }}

        function getEndDate(startIso, duration) {{
            return addDays(startIso, Math.max(1, Number(duration || 1)) - 1);
        }}

        function getCountryDateRange(country) {{
            if (country === 'all') {{
                if (!rows.length) {{
                    return {{ from: defaultFrom, to: defaultTo }};
                }}

                let minDate = rows[0].date;
                let maxDate = getEndDate(rows[0].date, rows[0].duration);

                rows.forEach((r) => {{
                    const start = r.date;
                    const end = getEndDate(r.date, r.duration);
                    if (start < minDate) {{
                        minDate = start;
                    }}
                    if (end > maxDate) {{
                        maxDate = end;
                    }}
                }});

                return {{ from: minDate, to: maxDate }};
            }}

            const countryRows = rows.filter((r) => r.country === country);
            if (!countryRows.length) {{
                return null;
            }}

            let minDate = countryRows[0].date;
            let maxDate = getEndDate(countryRows[0].date, countryRows[0].duration);

            countryRows.forEach((r) => {{
                const start = r.date;
                const end = getEndDate(r.date, r.duration);
                if (start < minDate) {{
                    minDate = start;
                }}
                if (end > maxDate) {{
                    maxDate = end;
                }}
            }});

            return {{ from: minDate, to: maxDate }};
        }}

        function applyCountryDateRange(country) {{
            const range = getCountryDateRange(country);
            if (!range) {{
                return;
            }}

            fromInput.value = range.from;
            toInput.value = range.to;

            if (fromInput.value > toInput.value) {{
                toInput.value = fromInput.value;
            }}
        }}

        function updateFullscreenButtonLabel() {{
            if (!toggleFullscreenBtn) {{
                return;
            }}
            toggleFullscreenBtn.textContent = document.fullscreenElement ? 'Exit fullscreen' : 'Maximize';
        }}

        async function toggleFullscreenView() {{
            if (!document.fullscreenElement) {{
                await document.documentElement.requestFullscreen();
            }} else {{
                await document.exitFullscreen();
            }}
            updateFullscreenButtonLabel();
        }}

        function closeNoteModal() {{
            if (!noteModalBackdrop) {{
                return;
            }}
            noteModalBackdrop.classList.remove('open');
            noteModalBackdrop.setAttribute('aria-hidden', 'true');
            activeNoteRow = null;
            activeNoteBadgeUpdater = null;
        }}

        function openNoteModal(rowRef, badgeUpdater) {{
            if (!noteModalBackdrop || !noteModalInput || !noteModalTitle) {{
                return;
            }}
            activeNoteRow = rowRef;
            activeNoteBadgeUpdater = badgeUpdater;
            noteModalTitle.textContent = `Note - ${{buildServerLabel(rowRef)}}`;
            if (noteRhelOk) {{
                noteRhelOk.checked = Boolean(rowRef.rhelOk);
            }}
            if (noteImeOk) {{
                noteImeOk.checked = Boolean(rowRef.imeOk);
            }}
            if (noteAppTestOk) {{
                noteAppTestOk.checked = Boolean(rowRef.appTestOk);
            }}
            if (noteRegressionTestOk) {{
                noteRegressionTestOk.checked = Boolean(rowRef.regressionTestOk);
            }}
            if (noteIpSwapOk) {{
                noteIpSwapOk.checked = Boolean(rowRef.ipSwapOk);
            }}
            if (noteIpSwapStartDate) {{
                noteIpSwapStartDate.value = String(rowRef.ipSwapStartDate || '');
            }}
            if (noteIpSwapEndDate) {{
                noteIpSwapEndDate.value = String(rowRef.ipSwapEndDate || rowRef.ipSwapStartDate || '');
            }}
            if (noteIpSwapFrom) {{
                noteIpSwapFrom.value = String(rowRef.ipSwapFrom || '');
            }}
            if (noteIpSwapTo) {{
                noteIpSwapTo.value = String(rowRef.ipSwapTo || '');
            }}
            if (notePatchOk) {{
                notePatchOk.checked = Boolean(rowRef.patchOk);
            }}
            if (notePatchNumber) {{
                notePatchNumber.value = String(rowRef.patchNumber || '');
            }}
            noteModalInput.value = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '');
            noteModalBackdrop.classList.add('open');
            noteModalBackdrop.setAttribute('aria-hidden', 'false');
            noteModalInput.focus();
            noteModalInput.select();
        }}

        function renderDragEditor(visibleRows, fromDate, toDate) {{
            if (!dragEditor) {{
                return;
            }}

            dragEditor.innerHTML = '';
            if (!visibleRows.length) {{
                const empty = document.createElement('div');
                empty.className = 'editor-empty';
                empty.textContent = 'No servers in the selected range.';
                dragEditor.appendChild(empty);
                return;
            }}

            const totalDays = Math.max(1, daysBetween(fromDate, toDate) + 1);
            const pxPerDay = 18;

            const expandDateWindow = (startIso, endIso) => {{
                if (fromInput && startIso && (!fromInput.value || startIso < fromInput.value)) {{
                    fromInput.value = startIso;
                }}
                if (toInput && endIso && (!toInput.value || endIso > toInput.value)) {{
                    toInput.value = endIso;
                }}
            }};

            const placeBar = (rowRef, barEl) => {{
                rowRef.duration = Math.max(1, Number(rowRef.duration || 1));
                const startOffset = Math.max(0, Math.min(totalDays - 1, daysBetween(fromDate, rowRef.date)));
                barEl.style.left = `${{startOffset * pxPerDay}}px`;
                barEl.style.width = `${{Math.max(12, rowRef.duration * pxPerDay)}}px`;
                barEl.style.background = lifecycleGradient(rowRef);
                barEl.title = lifecycleHoverText(rowRef);
            }};

            visibleRows.forEach((rowRef) => {{
                const rowEl = document.createElement('div');
                rowEl.className = 'editor-row';

                const nameEl = document.createElement('div');
                nameEl.className = 'editor-name';
                nameEl.textContent = buildServerLabel(rowRef);

                const nameWrapEl = document.createElement('div');
                nameWrapEl.className = 'editor-name-wrap';
                nameWrapEl.appendChild(nameEl);

                const noteBadge = document.createElement('span');
                noteBadge.className = 'note-badge';
                noteBadge.textContent = 'N';
                noteBadge.title = 'Server has a note';

                const updateNoteBadge = () => {{
                    const noteText = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '').trim();
                    const lifecycleInfo = [
                        rowRef.rhelOk ? 'RHEL9 ok' : '',
                        rowRef.imeOk ? 'IME V11 ok' : '',
                        rowRef.patchOk ? 'Patching ok' : '',
                        rowRef.patchNumber ? `Patch: ${{rowRef.patchNumber}}` : '',
                    ].filter(Boolean).join(' | ');
                    const hasMeta = Boolean(noteText || lifecycleInfo);
                    noteBadge.style.display = hasMeta ? 'inline-flex' : 'none';
                    noteBadge.title = [lifecycleInfo, noteText ? `Note: ${{noteText}}` : ''].filter(Boolean).join(' | ') || 'Server has a note';
                }};
                updateNoteBadge();
                nameWrapEl.appendChild(noteBadge);

                const noteBtn = document.createElement('button');
                noteBtn.className = 'editor-mini';
                noteBtn.type = 'button';
                noteBtn.textContent = 'Note';
                noteBtn.title = 'Edit note in row';
                nameWrapEl.appendChild(noteBtn);

                const trackEl = document.createElement('div');
                trackEl.className = 'editor-track';
                trackEl.style.width = `${{totalDays * pxPerDay}}px`;

                const barEl = document.createElement('div');
                barEl.className = 'editor-bar';
                placeBar(rowRef, barEl);

                const liveEl = document.createElement('span');
                liveEl.className = 'editor-live';
                liveEl.textContent = `${{rowRef.date}} -> ${{getEndDate(rowRef.date, rowRef.duration)}}`;
                barEl.appendChild(liveEl);

                const pickerEl = document.createElement('div');
                pickerEl.className = 'editor-picker';

                const fromLbl = document.createElement('label');
                fromLbl.textContent = 'From';
                const fromPicker = document.createElement('input');
                fromPicker.type = 'date';
                fromPicker.value = rowRef.date;
                fromLbl.appendChild(fromPicker);

                const toLbl = document.createElement('label');
                toLbl.textContent = 'To';
                const toPicker = document.createElement('input');
                toPicker.type = 'date';
                toPicker.value = getEndDate(rowRef.date, rowRef.duration);
                toLbl.appendChild(toPicker);

                const applyPickerBtn = document.createElement('button');
                applyPickerBtn.className = 'editor-mini';
                applyPickerBtn.type = 'button';
                applyPickerBtn.textContent = 'Set dates';

                const closePickerBtn = document.createElement('button');
                closePickerBtn.className = 'editor-mini';
                closePickerBtn.type = 'button';
                closePickerBtn.textContent = 'Close';

                pickerEl.appendChild(fromLbl);
                pickerEl.appendChild(toLbl);
                pickerEl.appendChild(applyPickerBtn);
                pickerEl.appendChild(closePickerBtn);

                const pickerHost = document.createElement('div');
                pickerHost.className = 'editor-picker-host';
                pickerHost.appendChild(pickerEl);

                trackEl.appendChild(barEl);
                rowEl.appendChild(nameWrapEl);
                rowEl.appendChild(trackEl);
                rowEl.appendChild(pickerHost);
                dragEditor.appendChild(rowEl);

                const openPicker = () => {{
                    fromPicker.value = rowRef.date;
                    toPicker.value = getEndDate(rowRef.date, rowRef.duration);
                    pickerEl.classList.add('open');
                }};

                const closePicker = () => {{
                    pickerEl.classList.remove('open');
                }};

                const startDrag = (handleNode, openPickerOnClick) => {{
                    handleNode.addEventListener('pointerdown', (ev) => {{
                        ev.preventDefault();
                        const dragStartX = ev.clientX;
                        const dragStartDate = rowRef.date;
                        let dragDays = 0;
                        let moved = false;

                        handleNode.style.cursor = 'grabbing';
                        barEl.style.cursor = 'grabbing';
                        liveEl.classList.add('open');

                        const onMove = (moveEv) => {{
                            const deltaPx = moveEv.clientX - dragStartX;
                            moved = moved || Math.abs(deltaPx) >= 3;
                            dragDays = Math.round(deltaPx / pxPerDay);
                            const previewDate = addDays(dragStartDate, dragDays);
                            rowRef.date = previewDate;
                            placeBar(rowRef, barEl);
                            liveEl.textContent = `${{previewDate}} -> ${{getEndDate(previewDate, rowRef.duration)}}`;
                            statusText.classList.remove('error');
                            statusText.textContent = `Moving: ${{previewDate}} -> ${{getEndDate(previewDate, rowRef.duration)}}`;
                        }};

                        const onEnd = () => {{
                            document.removeEventListener('pointermove', onMove);
                            document.removeEventListener('pointerup', onEnd);
                            document.removeEventListener('pointercancel', onCancel);
                            handleNode.style.cursor = 'grab';
                            barEl.style.cursor = 'grab';
                            liveEl.classList.remove('open');

                            if (!moved && openPickerOnClick) {{
                                openPicker();
                                return;
                            }}

                            pushUndoSnapshot();
                            rowRef.date = addDays(dragStartDate, dragDays);
                            expandDateWindow(rowRef.date, getEndDate(rowRef.date, rowRef.duration));
                            placeBar(rowRef, barEl);
                            statusText.classList.remove('error');
                            statusText.textContent = '';
                            renderChart();
                        }};

                        const onCancel = () => {{
                            document.removeEventListener('pointermove', onMove);
                            document.removeEventListener('pointerup', onEnd);
                            document.removeEventListener('pointercancel', onCancel);
                            rowRef.date = dragStartDate;
                            placeBar(rowRef, barEl);
                            handleNode.style.cursor = 'grab';
                            barEl.style.cursor = 'grab';
                            liveEl.classList.remove('open');
                            statusText.classList.remove('error');
                            statusText.textContent = '';
                            renderChart();
                        }};

                        document.addEventListener('pointermove', onMove);
                        document.addEventListener('pointerup', onEnd);
                        document.addEventListener('pointercancel', onCancel);
                    }});
                }};

                applyPickerBtn.addEventListener('click', () => {{
                    const pickedFrom = fromPicker.value;
                    const pickedTo = toPicker.value;
                    if (!pickedFrom || !pickedTo || pickedFrom > pickedTo) {{
                        statusText.classList.add('error');
                        statusText.textContent = 'Invalid date selection in the calendar.';
                        return;
                    }}

                    rowRef.date = pickedFrom;
                    if (pickedTo < rowRef.date) {{
                        statusText.classList.add('error');
                        statusText.textContent = 'To cannot be earlier than From.';
                        return;
                    }}

                    pushUndoSnapshot();
                    rowRef.duration = Math.max(1, daysBetween(rowRef.date, pickedTo) + 1);
                    expandDateWindow(rowRef.date, pickedTo);
                    placeBar(rowRef, barEl);
                    closePicker();
                    statusText.classList.remove('error');
                    statusText.textContent = '';
                    renderChart();
                }});

                closePickerBtn.addEventListener('click', closePicker);
                noteBtn.addEventListener('click', () => {{
                    openNoteModal(rowRef, updateNoteBadge);
                }});
                startDrag(nameEl, false);
                startDrag(barEl, true);
            }});
        }}

        function estimateDayWidth(svg) {{
            const widths = Array.from(svg.querySelectorAll('rect'))
                .map((r) => r.getBBox().width)
                .filter((w) => Number.isFinite(w) && w >= 10 && w <= 160)
                .sort((a, b) => a - b);

            if (!widths.length) {{
                return 24;
            }}
            return widths[0];
        }}

        function getVisibleRows(country, fromDate, toDate) {{
            const intersectsRange = (rangeStart, rangeEnd, filterStart, filterEnd) => {{
                if (!rangeStart && !rangeEnd) {{
                    return false;
                }}
                const start = String(rangeStart || rangeEnd || '');
                const end = String(rangeEnd || rangeStart || '');
                return start <= filterEnd && end >= filterStart;
            }};

            return rows
                .filter((r) => (country === 'all' ? true : r.country === country))
                .filter((r) => (
                    intersectsRange(r.date, getEndDate(r.date, Math.max(1, Number(r.duration || 1))), fromDate, toDate)
                    || intersectsRange(r.ipSwapStartDate, r.ipSwapEndDate, fromDate, toDate)
                ));
        }}

        function applyDragEditorFilters(sourceRows) {{
            const countryFilter = String(dragFilterCountry?.value || 'all');
            const serverFilter = String(dragFilterServer?.value || 'all').trim();
            const environmentFilter = String(dragFilterEnvironment?.value || 'all').trim();
            const fromFilter = String(dragFilterFrom?.value || '').trim();
            const toFilter = String(dragFilterTo?.value || '').trim();

            const intersectsRange = (rangeStart, rangeEnd, filterStart, filterEnd) => {{
                if (!rangeStart && !rangeEnd) {{
                    return false;
                }}
                const start = String(rangeStart || rangeEnd || '');
                const end = String(rangeEnd || rangeStart || '');
                return start <= filterEnd && end >= filterStart;
            }};

            return sourceRows.filter((rowRef) => {{
                if (countryFilter !== 'all' && rowRef.country !== countryFilter) {{
                    return false;
                }}

                if (serverFilter !== 'all' && String(rowRef.server || '').trim() !== serverFilter) {{
                    return false;
                }}

                if (environmentFilter !== 'all' && String(rowRef.environment || '').trim() !== environmentFilter) {{
                    return false;
                }}

                if (fromFilter || toFilter) {{
                    // Fix: Properly handle fromFilter and toFilter
                    // If only one is provided, use it as both start and end
                    const filterStart = fromFilter && toFilter ? Math.min(fromFilter, toFilter) : (fromFilter || toFilter || '');
                    const filterEnd = fromFilter && toFilter ? Math.max(fromFilter, toFilter) : (fromFilter || toFilter || '');
                    
                    if (filterStart || filterEnd) {{
                        const inMainRange = intersectsRange(
                            rowRef.date,
                            getEndDate(rowRef.date, Math.max(1, Number(rowRef.duration || 1))),
                            filterStart,
                            filterEnd,
                        );
                        const inSwapRange = intersectsRange(
                            rowRef.ipSwapStartDate,
                            rowRef.ipSwapEndDate,
                            filterStart,
                            filterEnd,
                        );
                        if (!inMainRange && !inSwapRange) {{
                            return false;
                        }}
                    }}
                }}

                return true;
            }});
        }}

        function updateOverviewStats() {{
            if (!overviewTotal || !overviewDone || !overviewPartial || !overviewRemaining) {{
                return;
            }}

            let done = 0;
            let partial = 0;
            let remaining = 0;
            rows.forEach((r) => {{
                const status = normalizeProgressStatus(String(r.status || 'none'));
                if (status === 'done') {{
                    done += 1;
                }} else if (status === 'partial') {{
                    partial += 1;
                }} else {{
                    remaining += 1;
                }}
            }});

            overviewTotal.textContent = String(rows.length);
            overviewDone.textContent = String(done);
            overviewPartial.textContent = String(partial);
            overviewRemaining.textContent = String(remaining);
        }}

        function getOverviewRows(kind) {{
            if (kind === 'total') {{
                return [...rows];
            }}
            return rows.filter((r) => {{
                const status = normalizeProgressStatus(String(r.status || 'none'));
                if (kind === 'done') {{
                    return status === 'done';
                }}
                if (kind === 'partial') {{
                    return status === 'partial';
                }}
                return status === 'none';
            }});
        }}

        function openOverviewModal(kind) {{
            if (!overviewModalBackdrop || !overviewModalTitle || !overviewModalList || !overviewModalEmpty) {{
                return;
            }}

            const labelMap = {{
                total: 'Total',
                done: 'Done',
                partial: 'Partial',
                remaining: 'Remaining',
            }};
            const selectedRows = getOverviewRows(kind).sort((a, b) => {{
                if (a.country !== b.country) {{
                    return a.country.localeCompare(b.country);
                }}
                if (a.date !== b.date) {{
                    return a.date.localeCompare(b.date);
                }}
                return buildServerLabel(a).localeCompare(buildServerLabel(b));
            }});

            overviewModalTitle.textContent = `Servers - ${{labelMap[kind] || 'Selection'}} (${{selectedRows.length}})`;
            overviewModalList.innerHTML = '';

            if (!selectedRows.length) {{
                overviewModalEmpty.hidden = false;
            }} else {{
                overviewModalEmpty.hidden = true;
                selectedRows.forEach((rowRef) => {{
                    const item = document.createElement('li');
                    const endDate = getEndDate(rowRef.date, Math.max(1, Number(rowRef.duration || 1)));
                    item.textContent = `${{rowRef.country}} | ${{buildServerLabel(rowRef)}} | ${{rowRef.date}} -> ${{endDate}}`;
                    overviewModalList.appendChild(item);
                }});
            }}

            overviewModalBackdrop.classList.add('open');
            overviewModalBackdrop.setAttribute('aria-hidden', 'false');
            overviewModalClose?.focus();
        }}

        function closeOverviewModal() {{
            if (!overviewModalBackdrop) {{
                return;
            }}
            overviewModalBackdrop.classList.remove('open');
            overviewModalBackdrop.setAttribute('aria-hidden', 'true');
        }}

        function getLatestPlanEnd() {{
            if (!rows.length) {{
                return toInput.value;
            }}
            return rows.reduce((latest, r) => {{
                const end = getEndDate(r.date, r.duration);
                return end > latest ? end : latest;
            }}, rows[0].date);
        }}

        function ensureDateBoundsCoverRows() {{
            const latestEnd = getLatestPlanEnd();
            if (toInput.value < latestEnd) {{
                toInput.value = latestEnd;
            }}
        }}

        function suggestPlanForAllServers() {{
            const start = planStartDate?.value || fromInput.value;
            const perDay = Math.max(1, Math.min(20, Number(planPerDay?.value || 2)));
            const skipWeekends = Boolean(planSkipWeekends?.checked);

            statusText.classList.remove('error');
            statusText.textContent = 'Calculating plan suggestion...';

            if (!start) {{
                statusText.classList.add('error');
                statusText.textContent = 'Enter a valid start date for the plan suggestion.';
                return;
            }}

            const ordered = [...rows].sort((a, b) => {{
                const ap = Number.isFinite(Number(a.priority)) ? Number(a.priority) : 9999;
                const bp = Number.isFinite(Number(b.priority)) ? Number(b.priority) : 9999;
                if (ap !== bp) {{
                    return ap - bp;
                }}
                if (a.country !== b.country) {{
                    return a.country.localeCompare(b.country);
                }}
                return a.server.localeCompare(b.server);
            }});

            let slotDate = nextAllowedDay(start, skipWeekends);
            let slotsUsed = 0;
            pushUndoSnapshot();

            ordered.forEach((rowRef) => {{
                if (slotsUsed >= perDay) {{
                    slotDate = nextAllowedDay(addDays(slotDate, 1), skipWeekends);
                    slotsUsed = 0;
                }}
                rowRef.date = slotDate;
                rowRef.duration = Math.max(1, Number(rowRef.duration || 1));
                slotsUsed += 1;
            }});

            ensureDateBoundsCoverRows();
            const latestEnd = getLatestPlanEnd();

            statusText.classList.remove('error');
            statusText.textContent = `Plan suggestion ready: ${{ordered.length}} servers, ${{perDay}} per day, range ${{start}} -> ${{latestEnd}}.`;
            renderChart();
        }}

        function renderFallbackChart(visibleRows, fromDate, toDate) {{
            const totalDays = Math.max(1, daysBetween(fromDate, toDate) + 1);
            const pxPerDay = 16;

            const root = document.createElement('div');
            root.className = 'fallback';
            const grid = document.createElement('div');
            grid.className = 'fallback-grid';

            visibleRows.forEach((rowRef) => {{
                const rowEl = document.createElement('div');
                rowEl.className = 'fallback-row';

                const nameEl = document.createElement('div');
                nameEl.className = 'fallback-name';
                nameEl.textContent = buildServerLabel(rowRef);

                const nameWrapEl = document.createElement('div');
                nameWrapEl.className = 'fallback-name-wrap';
                nameWrapEl.appendChild(nameEl);

                const noteBadge = document.createElement('span');
                noteBadge.className = 'note-badge';
                noteBadge.textContent = 'N';
                const noteTextForBadge = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '').trim();
                if (noteTextForBadge) {{
                    noteBadge.title = `Note: ${{noteTextForBadge}}`;
                    nameWrapEl.appendChild(noteBadge);
                }}

                const trackEl = document.createElement('div');
                trackEl.className = 'fallback-track';
                trackEl.style.width = `${{totalDays * pxPerDay}}px`;

                const barEl = document.createElement('div');
                barEl.className = 'fallback-bar';
                const startOffset = Math.max(0, Math.min(totalDays - 1, daysBetween(fromDate, rowRef.date)));
                const duration = Math.max(1, Number(rowRef.duration || 1));
                barEl.style.left = `${{startOffset * pxPerDay}}px`;
                barEl.style.width = `${{Math.max(8, duration * pxPerDay)}}px`;
                barEl.style.background = lifecycleGradient(rowRef);
                barEl.title = lifecycleHoverText(rowRef);

                trackEl.appendChild(barEl);
                if (rowRef.ipSwapStartDate || rowRef.ipSwapEndDate) {{
                    const swapStart = String(rowRef.ipSwapStartDate || rowRef.ipSwapEndDate || '');
                    const swapEnd = String(rowRef.ipSwapEndDate || rowRef.ipSwapStartDate || '');
                    if (swapStart && swapEnd) {{
                        const swapBarEl = document.createElement('div');
                        swapBarEl.className = 'fallback-bar';
                        const swapOffset = Math.max(0, Math.min(totalDays - 1, daysBetween(fromDate, swapStart)));
                        const swapDuration = Math.max(1, daysBetween(swapStart, swapEnd) + 1);
                        swapBarEl.style.left = `${{swapOffset * pxPerDay}}px`;
                        swapBarEl.style.width = `${{Math.max(8, swapDuration * pxPerDay)}}px`;
                        swapBarEl.style.top = '20px';
                        swapBarEl.style.height = '10px';
                        swapBarEl.style.borderRadius = '5px';
                        swapBarEl.style.background = rowRef.ipSwapOk ? '#16a34a' : '#f97316';
                        swapBarEl.title = `IP Swap plan\n${{lifecycleHoverText(rowRef)}}`;
                        trackEl.appendChild(swapBarEl);
                        trackEl.style.minHeight = '34px';
                    }}
                }}
                rowEl.appendChild(nameWrapEl);
                rowEl.appendChild(trackEl);
                grid.appendChild(rowEl);
            }});

            const datesEl = document.createElement('div');
            datesEl.className = 'fallback-dates';
            datesEl.textContent = `Showing fallback chart for ${{fromDate}} -> ${{toDate}}.`;

            root.appendChild(grid);
            root.appendChild(datesEl);
            chart.innerHTML = '';
            chart.appendChild(root);
        }}

        function addServerTooltips(visibleRows) {{
            const svg = chart.querySelector('svg');
            if (!svg || !visibleRows.length) {{
                return;
            }}

            const ensureHoverTip = () => {{
                let tip = document.getElementById('serverHoverTip');
                if (tip) {{
                    return tip;
                }}
                tip = document.createElement('div');
                tip.id = 'serverHoverTip';
                tip.style.position = 'fixed';
                tip.style.zIndex = '9999';
                tip.style.pointerEvents = 'none';
                tip.style.background = '#0f172a';
                tip.style.color = '#fff';
                tip.style.fontSize = '12px';
                tip.style.borderRadius = '6px';
                tip.style.padding = '4px 8px';
                tip.style.boxShadow = '0 2px 10px rgba(15, 23, 42, 0.25)';
                tip.style.whiteSpace = 'pre-line';
                tip.style.maxWidth = '360px';
                tip.style.display = 'none';
                document.body.appendChild(tip);
                return tip;
            }};
            const hoverTip = ensureHoverTip();

            const removeDirectTitle = (node) => {{
                if (!node || !node.children) {{
                    return;
                }}
                const titleChild = Array.from(node.children).find(
                    (ch) => String(ch.tagName || '').toLowerCase() === 'title'
                );
                if (titleChild) {{
                    titleChild.remove();
                }}
            }};

            const rowsByLabel = new Map();
            visibleRows.forEach((r) => {{
                const baseKey = normalizeLabel(buildServerLabel(r));
                const swapKey = normalizeLabel(`${{buildServerLabel(r)}} [IP Swap]`);
                const baseList = rowsByLabel.get(baseKey) || [];
                baseList.push(r);
                rowsByLabel.set(baseKey, baseList);
                const swapList = rowsByLabel.get(swapKey) || [];
                swapList.push(r);
                rowsByLabel.set(swapKey, swapList);
            }});

            const cursor = new Map();
            svg.querySelectorAll('text').forEach((node) => {{
                const key = normalizeLabel(node.textContent || '');
                if (!rowsByLabel.has(key)) {{
                    return;
                }}
                const idx = cursor.get(key) || 0;
                const rowsForLabel = rowsByLabel.get(key);
                if (!rowsForLabel || idx >= rowsForLabel.length) {{
                    return;
                }}
                const rowRef = rowsForLabel[idx];
                cursor.set(key, idx + 1);
                const tooltipText = lifecycleHoverText(rowRef);
                removeDirectTitle(node);
                node.style.cursor = 'help';
                node.style.pointerEvents = 'all';

                const showTip = (ev) => {{
                    hoverTip.textContent = tooltipText;
                    hoverTip.style.display = 'block';
                    hoverTip.style.left = `${{ev.clientX + 12}}px`;
                    hoverTip.style.top = `${{ev.clientY + 12}}px`;
                }};
                const moveTip = (ev) => {{
                    hoverTip.style.left = `${{ev.clientX + 12}}px`;
                    hoverTip.style.top = `${{ev.clientY + 12}}px`;
                }};
                const hideTip = () => {{
                    hoverTip.style.display = 'none';
                }};

                node.addEventListener('mouseenter', showTip);
                node.addEventListener('mousemove', moveTip);
                node.addEventListener('mouseleave', hideTip);
            }});
        }}

        function applyTaskStatusColors(visibleRows) {{
            const svg = chart.querySelector('svg');
            if (!svg || !visibleRows.length) {{
                return;
            }}

            const rowsByLabel = new Map();
            visibleRows.forEach((r) => {{
                const key = normalizeLabel(buildServerLabel(r));
                const list = rowsByLabel.get(key) || [];
                list.push(r);
                rowsByLabel.set(key, list);
            }});

            const taskRects = Array.from(svg.querySelectorAll('rect'))
                .filter((r) => {{
                    const box = r.getBBox();
                    return box.width >= 10 && box.height >= 8;
                }});

            function findTaskRectForLabel(labelNode) {{
                const labelBox = labelNode.getBBox();
                const labelMidY = labelBox.y + labelBox.height / 2;
                let bestRect = null;
                let bestScore = Number.POSITIVE_INFINITY;

                taskRects.forEach((rect) => {{
                    const box = rect.getBBox();
                    const rectMidY = box.y + box.height / 2;
                    const dy = Math.abs(rectMidY - labelMidY);
                    const dxPenalty = box.x < (labelBox.x + labelBox.width + 20) ? 40 : 0;
                    const score = dy + dxPenalty;
                    if (dy <= 12 && score < bestScore) {{
                        bestScore = score;
                        bestRect = rect;
                    }}
                }});

                return bestRect;
            }}

            let defs = svg.querySelector('defs');
            if (!defs) {{
                defs = document.createElementNS('http://www.w3.org/2000/svg', 'defs');
                svg.insertBefore(defs, svg.firstChild);
            }}

            const labelCursor = new Map();
            svg.querySelectorAll('text').forEach((node) => {{
                const key = normalizeLabel(node.textContent || '');
                if (!rowsByLabel.has(key)) {{
                    return;
                }}

                const idx = labelCursor.get(key) || 0;
                const rowsForLabel = rowsByLabel.get(key);
                if (!rowsForLabel || idx >= rowsForLabel.length) {{
                    return;
                }}
                const rowRef = rowsForLabel[idx];
                labelCursor.set(key, idx + 1);
                rowRef.status = normalizeProgressStatus(rowRef.status);

                const barNode = findTaskRectForLabel(node);
                if (!barNode) {{
                    return;
                }}

                if (rowRef.status === 'done') {{
                    barNode.style.fill = '#22c55e';
                }} else if (rowRef.status === 'partial') {{
                    const gradId = `status_partial_${{slug(rowRef.country)}}_${{idx}}`;
                    let grad = svg.querySelector(`#${{gradId}}`);
                    if (!grad) {{
                        grad = document.createElementNS('http://www.w3.org/2000/svg', 'linearGradient');
                        grad.setAttribute('id', gradId);
                        grad.setAttribute('x1', '0%');
                        grad.setAttribute('y1', '0%');
                        grad.setAttribute('x2', '100%');
                        grad.setAttribute('y2', '0%');
                        const stopA = document.createElementNS('http://www.w3.org/2000/svg', 'stop');
                        stopA.setAttribute('offset', '50%');
                        stopA.setAttribute('stop-color', '#22c55e');
                        const stopB = document.createElementNS('http://www.w3.org/2000/svg', 'stop');
                        stopB.setAttribute('offset', '50%');
                        stopB.setAttribute('stop-color', '#ef4444');
                        grad.appendChild(stopA);
                        grad.appendChild(stopB);
                        defs.appendChild(grad);
                    }}
                    barNode.style.fill = `url(#${{gradId}})`;
                }} else {{
                    barNode.style.fill = '#ef4444';
                }}
                barNode.style.stroke = '#475569';
                barNode.style.strokeWidth = '0.8';
                barNode.setAttribute('data-progress-status', rowRef.status);
            }});
        }}

        function addServerRow() {{
            if (!addCountry || !addServerName || !addFrom || !addDuration) {{
                return;
            }}

            const freeCountry = (addCountryFree?.value || '').trim();
            const country = freeCountry || addCountry.value;
            const serverName = (addServerName.value || '').trim();
            const env = (addEnv?.value || '').trim();
            const note = (addNote?.value || '').trim();
            const from = addFrom.value;
            const duration = Math.max(1, Math.min(60, Number(addDuration.value || 1)));

            if (!country || !serverName || !from) {{
                addServerMsg.classList.add('error');
                addServerMsg.textContent = 'Fill in country, server name, and from date.';
                return;
            }}

            pushUndoSnapshot();
            rows.push({{
                priority: rows.length ? Math.max(...rows.map((r) => Number(r.priority || 0))) + 1 : 1,
                country,
                server: serverName,
                environment: env,
                date: from,
                duration,
                status: 'none',
                rhelOk: false,
                imeOk: false,
                appTestOk: false,
                regressionTestOk: false,
                ipSwapOk: false,
                ipSwapStartDate: '',
                ipSwapEndDate: '',
                ipSwapFrom: '',
                ipSwapTo: '',
                patchOk: false,
                patchNumber: '',
                notesProd: note,
                notesTest: '',
                notes: note,
            }});

            if (!countries.includes(country)) {{
                countries.push(country);
                createCountryButton(country);
                refreshCountrySelect();
            }}

            addServerMsg.classList.remove('error');
            addServerMsg.textContent = `Server added: ${{country}} / ${{serverName}}`;
            if (addCountryFree) {{
                addCountryFree.value = '';
            }}
            addServerName.value = '';
            addEnv.value = '';
            if (addNote) {{
                addNote.value = '';
            }}
            addDuration.value = '1';
            ensureDateBoundsCoverRows();
            renderChart();
            refreshCountrySelect();
        }}

        function removeSelectedServerRow() {{
            if (!removeServer || !removeCountry) {{
                return;
            }}

            const selectedIdx = Number(removeServer.value);
            if (!Number.isFinite(selectedIdx) || selectedIdx < 0 || selectedIdx >= removeCandidates.length) {{
                removeServerMsg.classList.add('error');
                removeServerMsg.textContent = 'Select a server to remove.';
                return;
            }}

            const rowRef = removeCandidates[selectedIdx];
            const removedCountry = rowRef.country;
            const rowIndex = rows.indexOf(rowRef);
            if (rowIndex < 0) {{
                return;
            }}

            pushUndoSnapshot();
            rows.splice(rowIndex, 1);

            if (!rows.some((r) => r.country === removedCountry)) {{
                const countryIndex = countries.indexOf(removedCountry);
                if (countryIndex >= 0) {{
                    countries.splice(countryIndex, 1);
                }}
                removeCountryButton(removedCountry);
                if (selectedCountry === removedCountry) {{
                    selectedCountry = countries.length ? countries[0] : 'all';
                    updateActiveButton();
                }}
            }}

            removeServerMsg.classList.remove('error');
            removeServerMsg.textContent = `Removed: ${{getServerLabel(rowRef)}}`;
            statusText.classList.remove('error');
            statusText.textContent = '';
            refreshCountrySelect();
            renderChart();
        }}

        function exportUpdatedExcel() {{
            if (!window.XLSX) {{
                statusText.classList.add('error');
                statusText.textContent = 'Could not load the Excel export library.';
                return;
            }}

            const rowsForExport = [...rows].sort((a, b) => {{
                if (a.country !== b.country) {{
                    return a.country.localeCompare(b.country);
                }}
                if (a.date !== b.date) {{
                    return a.date.localeCompare(b.date);
                }}
                return a.server.localeCompare(b.server);
            }});

            const header = ['Priority', 'Country', 'Server', 'Environment', 'Upgrade Date', 'From Date', 'To Date', 'Duration Days', 'RHEL9 OK', 'IME V11 OK', 'Application Test OK', 'Regression Test OK', 'IP Swap OK', 'IP Swap From Date', 'IP Swap To Date', 'IP Swap From', 'IP Swap To', 'Patching OK', 'Patch Number', 'Status', 'Notes Prod', 'Notes Test'];
            const aoa = [header];
            rowsForExport.forEach((r, idx) => {{
                const fromDate = r.date;
                const durationDays = Math.max(1, Number(r.duration || 1));
                const toDate = getEndDate(fromDate, durationDays);
                aoa.push([
                    idx + 1,
                    r.country,
                    r.server,
                    r.environment || '',
                    fromDate,
                    fromDate,
                    toDate,
                    durationDays,
                    r.rhelOk ? 'yes' : '',
                    r.imeOk ? 'yes' : '',
                    r.appTestOk ? 'yes' : '',
                    r.regressionTestOk ? 'yes' : '',
                    r.ipSwapOk ? 'yes' : '',
                    String(r.ipSwapStartDate || ''),
                    String(r.ipSwapEndDate || r.ipSwapStartDate || ''),
                    String(r.ipSwapFrom || ''),
                    String(r.ipSwapTo || ''),
                    r.patchOk ? 'yes' : '',
                    String(r.patchNumber || ''),
                    normalizeProgressStatus(String(r.status || 'none')),
                    String(r.notesProd || r.notes || ''),
                    String(r.notesTest || ''),
                ]);
            }});

            const ws = window.XLSX.utils.aoa_to_sheet(aoa);
            ws['!cols'] = [
                {{ wch: 10 }},
                {{ wch: 16 }},
                {{ wch: 44 }},
                {{ wch: 14 }},
                {{ wch: 14 }},
                {{ wch: 14 }},
                {{ wch: 14 }},
                {{ wch: 14 }},
                {{ wch: 10 }},
                {{ wch: 10 }},
                {{ wch: 16 }},
                {{ wch: 16 }},
                {{ wch: 14 }},
                {{ wch: 14 }},
                {{ wch: 24 }},
                {{ wch: 24 }},
                {{ wch: 12 }},
                {{ wch: 18 }},
                {{ wch: 10 }},
                {{ wch: 42 }},
                {{ wch: 42 }},
            ];

            const wb = window.XLSX.utils.book_new();
            window.XLSX.utils.book_append_sheet(wb, ws, 'Plan');

            const stamp = new Date().toISOString().slice(0, 10);
            const fileName = `upgrade_plan_updated_${{stamp}}.xlsx`;
            window.XLSX.writeFile(wb, fileName);
            statusText.classList.remove('error');
            statusText.textContent = `Excel exported with updated dates (${{rowsForExport.length}} rows).`;
        }}

        function confluenceCell(value) {{
            return String(value || '')
                .replace(/\|/g, '\\|')
                .replace(/\\r?\\n/g, '<br/>');
        }}

        function downloadTextFile(fileName, content) {{
            const blob = new Blob([content], {{ type: 'text/plain;charset=utf-8' }});
            const url = URL.createObjectURL(blob);
            const link = document.createElement('a');
            link.href = url;
            link.download = fileName;
            document.body.appendChild(link);
            link.click();
            link.remove();
            URL.revokeObjectURL(url);
        }}

        function exportConfluenceTable() {{
            const rowsForExport = [...rows].sort((a, b) => {{
                if (a.country !== b.country) {{
                    return a.country.localeCompare(b.country);
                }}
                if (a.date !== b.date) {{
                    return a.date.localeCompare(b.date);
                }}
                return a.server.localeCompare(b.server);
            }});

            const headers = [
                'Priority',
                'Country',
                'Server',
                'Environment',
                'From Date',
                'To Date',
                'Duration Days',
                'RHEL9 OK',
                'IME V11 OK',
                'Patching OK',
                'Patch Number',
                'Status',
                'Notes Prod',
                'Notes Test',
            ];

            const lines = [];
            lines.push(`||${{headers.map(confluenceCell).join('||')}}||`);

            rowsForExport.forEach((r, idx) => {{
                const fromDate = r.date;
                const durationDays = Math.max(1, Number(r.duration || 1));
                const toDate = getEndDate(fromDate, durationDays);
                const data = [
                    idx + 1,
                    r.country,
                    r.server,
                    r.environment || '',
                    fromDate,
                    toDate,
                    durationDays,
                    r.rhelOk ? 'yes' : 'no',
                    r.imeOk ? 'yes' : 'no',
                    r.patchOk ? 'yes' : 'no',
                    String(r.patchNumber || ''),
                    normalizeProgressStatus(String(r.status || 'none')),
                    String(r.notesProd || r.notes || ''),
                    String(r.notesTest || ''),
                ];
                lines.push(`|${{data.map(confluenceCell).join('|')}}|`);
            }});

            const stamp = new Date().toISOString().slice(0, 10);
            const fileName = `upgrade_plan_confluence_${{stamp}}.txt`;
            downloadTextFile(fileName, lines.join('\\n'));
            statusText.classList.remove('error');
            statusText.textContent = `Confluence export ready (${{rowsForExport.length}} rows).`;
        }}

        function exportConfluenceImage() {{
            const svg = chart.querySelector('svg');
            if (!svg) {{
                statusText.classList.add('error');
                statusText.textContent = 'No chart found to export as an image.';
                return;
            }}

            const serializer = new XMLSerializer();
            let source = serializer.serializeToString(svg);
            if (!source.includes('xmlns="http://www.w3.org/2000/svg"')) {{
                source = source.replace('<svg', '<svg xmlns="http://www.w3.org/2000/svg"');
            }}
            if (!source.includes('xmlns:xlink="http://www.w3.org/1999/xlink"')) {{
                source = source.replace('<svg', '<svg xmlns:xlink="http://www.w3.org/1999/xlink"');
            }}

            const viewBox = svg.viewBox?.baseVal;
            const fallbackRect = svg.getBoundingClientRect();
            const width = Math.max(1, Math.ceil(viewBox?.width || fallbackRect.width || 1600));
            const height = Math.max(1, Math.ceil(viewBox?.height || fallbackRect.height || 900));

            const scale = 2;
            const canvas = document.createElement('canvas');
            canvas.width = width * scale;
            canvas.height = height * scale;
            const ctx = canvas.getContext('2d');
            if (!ctx) {{
                statusText.classList.add('error');
                statusText.textContent = 'Could not create image canvas for export.';
                return;
            }}

            const svgBlob = new Blob([source], {{ type: 'image/svg+xml;charset=utf-8' }});
            const svgUrl = URL.createObjectURL(svgBlob);
            const img = new Image();

            img.onload = () => {{
                ctx.setTransform(scale, 0, 0, scale, 0, 0);
                ctx.fillStyle = '#ffffff';
                ctx.fillRect(0, 0, width, height);
                ctx.drawImage(img, 0, 0, width, height);
                URL.revokeObjectURL(svgUrl);

                canvas.toBlob((blob) => {{
                    if (!blob) {{
                        statusText.classList.add('error');
                        statusText.textContent = 'Could not create PNG file for Confluence export.';
                        return;
                    }}
                    const url = URL.createObjectURL(blob);
                    const link = document.createElement('a');
                    const stamp = new Date().toISOString().slice(0, 10);
                    link.href = url;
                    link.download = `upgrade_plan_confluence_gantt_${{stamp}}.png`;
                    document.body.appendChild(link);
                    link.click();
                    link.remove();
                    URL.revokeObjectURL(url);
                    statusText.classList.remove('error');
                    statusText.textContent = 'Confluence image exported (PNG).';
                }}, 'image/png');
            }};

            img.onerror = () => {{
                URL.revokeObjectURL(svgUrl);
                statusText.classList.add('error');
                statusText.textContent = 'Could not read SVG for image export.';
            }};

            img.src = svgUrl;
        }}

        function addDragToReschedule(visibleRows) {{
            const svg = chart.querySelector('svg');
            if (!svg) {{
                return;
            }}
            svg.style.touchAction = 'none';

            const ensureDragLiveTip = () => {{
                let tip = document.getElementById('dragLiveTip');
                if (tip) {{
                    return tip;
                }}
                tip = document.createElement('div');
                tip.id = 'dragLiveTip';
                tip.style.position = 'fixed';
                tip.style.zIndex = '10000';
                tip.style.pointerEvents = 'none';
                tip.style.background = '#0f172a';
                tip.style.color = '#fff';
                tip.style.fontSize = '12px';
                tip.style.borderRadius = '6px';
                tip.style.padding = '4px 8px';
                tip.style.boxShadow = '0 2px 10px rgba(15, 23, 42, 0.25)';
                tip.style.display = 'none';
                document.body.appendChild(tip);
                return tip;
            }};

            const dragLiveTip = ensureDragLiveTip();

            const dayWidth = estimateDayWidth(svg);
            const rowsByLabel = new Map();
            visibleRows.forEach((r) => {{
                const label = normalizeLabel(buildServerLabel(r));
                const arr = rowsByLabel.get(label) || [];
                arr.push(r);
                rowsByLabel.set(label, arr);
            }});

            const taskRects = Array.from(svg.querySelectorAll('rect'))
                .filter((r) => {{
                    const box = r.getBBox();
                    return box.width >= 10 && box.height >= 8;
                }});

            function findTaskRectForLabel(labelNode) {{
                const labelBox = labelNode.getBBox();
                const labelMidY = labelBox.y + labelBox.height / 2;
                let bestRect = null;
                let bestScore = Number.POSITIVE_INFINITY;

                taskRects.forEach((rect) => {{
                    const box = rect.getBBox();
                    const rectMidY = box.y + box.height / 2;
                    const dy = Math.abs(rectMidY - labelMidY);
                    const dxPenalty = box.x < (labelBox.x + labelBox.width + 20) ? 40 : 0;
                    const score = dy + dxPenalty;
                    if (dy <= 12 && score < bestScore) {{
                        bestScore = score;
                        bestRect = rect;
                    }}
                }});

                return bestRect;
            }}

            function applyTranslate(node, deltaPx) {{
                if (!node) {{
                    return;
                }}
                const base = node.getAttribute('data-base-transform') || '';
                node.setAttribute('transform', `${{base}} translate(${{deltaPx}}, 0)`.trim());
            }}

            function clearTranslate(node) {{
                if (!node) {{
                    return;
                }}
                const base = node.getAttribute('data-base-transform') || '';
                if (base) {{
                    node.setAttribute('transform', base);
                }} else {{
                    node.removeAttribute('transform');
                }}
            }}

            function bindDrag(handleNode, labelNode, barNode, rowRef) {{
                if (!handleNode) {{
                    return;
                }}

                if (!labelNode.getAttribute('data-base-transform')) {{
                    labelNode.setAttribute('data-base-transform', labelNode.getAttribute('transform') || '');
                }}
                if (barNode && !barNode.getAttribute('data-base-transform')) {{
                    barNode.setAttribute('data-base-transform', barNode.getAttribute('transform') || '');
                }}

                handleNode.style.cursor = 'grab';
                handleNode.style.pointerEvents = 'all';
                labelNode.style.pointerEvents = 'all';
                if (barNode) {{
                    barNode.style.pointerEvents = 'all';
                }}
                handleNode.addEventListener('pointerdown', (ev) => {{
                    ev.preventDefault();
                    const startX = ev.clientX;
                    const originalDate = rowRef.date;
                    const originalDuration = Math.max(1, Number(rowRef.duration || 1));
                    let previewDays = 0;

                    handleNode.style.cursor = 'grabbing';
                    labelNode.style.cursor = 'grabbing';
                    if (barNode) {{
                        barNode.style.cursor = 'grabbing';
                    }}

                    const onPointerMove = (moveEv) => {{
                        const deltaPx = moveEv.clientX - startX;
                        previewDays = Math.round(deltaPx / dayWidth);
                        const snappedPx = previewDays * dayWidth;
                        const previewStart = addDays(originalDate, previewDays);
                        const previewEnd = getEndDate(previewStart, originalDuration);

                        applyTranslate(labelNode, snappedPx);
                        applyTranslate(barNode, snappedPx);

                        dragLiveTip.textContent = `${{previewStart}} -> ${{previewEnd}}`;
                        dragLiveTip.style.display = 'block';
                        dragLiveTip.style.left = `${{moveEv.clientX + 12}}px`;
                        dragLiveTip.style.top = `${{moveEv.clientY + 12}}px`;
                        statusText.classList.remove('error');
                        statusText.textContent = `Moving: ${{previewStart}} -> ${{previewEnd}}`;
                    }};

                    const onPointerUp = () => {{
                        document.removeEventListener('pointermove', onPointerMove);
                        document.removeEventListener('pointerup', onPointerUp);
                        document.removeEventListener('pointercancel', onPointerCancel);

                        clearTranslate(labelNode);
                        clearTranslate(barNode);
                        dragLiveTip.style.display = 'none';
                        statusText.classList.remove('error');
                        statusText.textContent = '';

                        handleNode.style.cursor = 'grab';
                        labelNode.style.cursor = 'grab';
                        if (barNode) {{
                            barNode.style.cursor = 'grab';
                        }}

                        if (!previewDays) {{
                            renderChart();
                            return;
                        }}

                        pushUndoSnapshot();
                        const shifted = addDays(originalDate, previewDays);
                        rowRef.date = shifted;
                        renderChart();
                    }};

                    const onPointerCancel = () => {{
                        document.removeEventListener('pointermove', onPointerMove);
                        document.removeEventListener('pointerup', onPointerUp);
                        document.removeEventListener('pointercancel', onPointerCancel);

                        clearTranslate(labelNode);
                        clearTranslate(barNode);
                        dragLiveTip.style.display = 'none';
                        statusText.classList.remove('error');
                        statusText.textContent = '';

                        handleNode.style.cursor = 'grab';
                        labelNode.style.cursor = 'grab';
                        if (barNode) {{
                            barNode.style.cursor = 'grab';
                        }}
                        renderChart();
                    }};

                    document.addEventListener('pointermove', onPointerMove);
                    document.addEventListener('pointerup', onPointerUp);
                    document.addEventListener('pointercancel', onPointerCancel);
                }});
            }}

            const labelCursor = new Map();
            svg.querySelectorAll('text').forEach((node) => {{
                const label = normalizeLabel(node.textContent || '');
                if (!rowsByLabel.has(label)) {{
                    return;
                }}

                const idx = labelCursor.get(label) || 0;
                const matches = rowsByLabel.get(label);
                if (!matches || idx >= matches.length) {{
                    return;
                }}
                const rowRef = matches[idx];
                labelCursor.set(label, idx + 1);

                const barNode = findTaskRectForLabel(node);
                bindDrag(node, node, barNode, rowRef);
                if (barNode) {{
                    bindDrag(barNode, node, barNode, rowRef);
                }}
            }});
        }}

        function buildMermaid(country, fromDate, toDate, visibleRowsOverride = null) {{
            const title = country === 'all' ? 'Upgrade Plan 2026' : `Upgrade Plan 2026 - ${{country}}`;
            const visibleRows = Array.isArray(visibleRowsOverride)
                ? visibleRowsOverride
                : getVisibleRows(country, fromDate, toDate);
            const planDays = Math.max(1, daysBetween(fromDate, toDate) + 1);

            const orderedCountries = country === 'all' ? countries : [country];

            const lines = [
                'gantt',
                `    title ${{title}}`,
                '    dateFormat  YYYY-MM-DD',
                '    axisFormat  %Y-%m-%d',
                '',
                '    section Tidsram',
                `    Startdatum                                  :milestone, frame_start, ${{fromDate}}, 1d`,
                `    Hela planperioden                           :frame_period, ${{fromDate}}, ${{planDays}}d`,
                `    Slutdatum                                   :milestone, frame_end, ${{toDate}}, 1d`,
                '',
            ];

            let hasTask = false;
            for (const c of orderedCountries) {{
                const countryRows = visibleRows.filter((r) => r.country === c);
                if (!countryRows.length) {{
                    continue;
                }}
                hasTask = true;
                lines.push(`    section ${{c}}`);
                countryRows.forEach((r, index) => {{
                    const label = sanitizeMermaidLabel(buildServerLabel(r));
                    const duration = Math.max(1, Number(r.duration || 1));
                    lines.push(`    ${{label}} :${{slug(c)}}_${{index + 1}}, ${{r.date}}, ${{duration}}d`);
                    const swapStart = String(r.ipSwapStartDate || r.ipSwapEndDate || '');
                    const swapEnd = String(r.ipSwapEndDate || r.ipSwapStartDate || '');
                    if (swapStart && swapEnd) {{
                        const swapDuration = Math.max(1, daysBetween(swapStart, swapEnd) + 1);
                        const swapLabel = sanitizeMermaidLabel(`${{buildServerLabel(r)}} [IP Swap]`);
                        const swapState = r.ipSwapOk ? 'done, ' : 'active, ';
                        lines.push(`    ${{swapLabel}} :${{swapState}}${{slug(c)}}_swap_${{index + 1}}, ${{swapStart}}, ${{swapDuration}}d`);
                    }}
                }});
                lines.push('');
            }}

            if (!hasTask) {{
                lines.push('    section Urval');
                lines.push(`    Inga servrar i valt urval                    :milestone, no_data, ${{fromDate}}, 1d`);
            }}

            return lines.join('\\n');
        }}

        async function renderChart() {{
            const fromDate = fromInput.value;
            const toDate = toInput.value;

            if (!fromDate || !toDate || fromDate > toDate) {{
                statusText.classList.add('error');
                statusText.textContent = 'Invalid date range. From must be less than or equal to To.';
                chart.innerHTML = '';
                return;
            }}

            statusText.classList.remove('error');
            statusText.textContent = '';
            updateOverviewStats();

            const baseRows = getVisibleRows(selectedCountry, fromDate, toDate);
            const visibleRows = applyDragEditorFilters(baseRows);
            renderDragEditor(visibleRows, fromDate, toDate);

            if (!mermaid || typeof mermaid.render !== 'function') {{
                statusText.classList.add('error');
                statusText.textContent = 'Could not load the Mermaid library. Showing fallback chart.';
                renderFallbackChart(visibleRows, fromDate, toDate);
                return;
            }}

            const mermaidText = buildMermaid(selectedCountry, fromDate, toDate, visibleRows);
            const renderId = `gantt_${{Date.now()}}`;
            try {{
                const {{ svg }} = await mermaid.render(renderId, mermaidText);
                chart.innerHTML = svg;
                addDragToReschedule(visibleRows);
                addServerTooltips(visibleRows);
            }} catch (err) {{
                statusText.classList.add('error');
                const detail = err && err.message ? ` (${{err.message}})` : '';
                statusText.textContent = `Could not render the chart. Showing fallback chart.${{detail}}`;
                renderFallbackChart(visibleRows, fromDate, toDate);
            }}
        }}

        buttons.forEach((btn) => {{
            btn.addEventListener('click', () => {{
                selectedCountry = btn.dataset.country;
                updateActiveButton();
                applyCountryDateRange(selectedCountry);
                refreshCountrySelect();
                renderChart();
            }});
        }});

        fromInput.addEventListener('change', renderChart);
        toInput.addEventListener('change', renderChart);
        resetBtn.addEventListener('click', () => {{
            selectedCountry = 'all';
            fromInput.value = defaultFrom;
            toInput.value = defaultTo;
            if (dragFilterCountry) {{
                dragFilterCountry.value = 'all';
            }}
            if (dragFilterServer) {{
                dragFilterServer.value = 'all';
            }}
            if (dragFilterEnvironment) {{
                dragFilterEnvironment.value = 'all';
            }}
            if (dragFilterFrom) {{
                dragFilterFrom.value = '';
            }}
            if (dragFilterTo) {{
                dragFilterTo.value = '';
            }}
            updateActiveButton();
            refreshCountrySelect();
            renderChart();
        }});

        const bindDragFilterInput = (node, eventName = 'change') => {{
            if (!node) {{
                return;
            }}
            node.addEventListener(eventName, renderChart);
        }};
        bindDragFilterInput(dragFilterCountry, 'change');
        bindDragFilterInput(dragFilterServer, 'change');
        bindDragFilterInput(dragFilterEnvironment, 'change');
        bindDragFilterInput(dragFilterFrom, 'change');
        bindDragFilterInput(dragFilterTo, 'change');
        dragFilterReset?.addEventListener('click', () => {{
            if (dragFilterCountry) {{
                dragFilterCountry.value = 'all';
            }}
            if (dragFilterServer) {{
                dragFilterServer.value = 'all';
            }}
            if (dragFilterEnvironment) {{
                dragFilterEnvironment.value = 'all';
            }}
            if (dragFilterFrom) {{
                dragFilterFrom.value = '';
            }}
            if (dragFilterTo) {{
                dragFilterTo.value = '';
            }}
            renderChart();
        }});

        addServerBtn?.addEventListener('click', addServerRow);
        saveChangesBtn?.addEventListener('click', saveCurrentState);
        undoBtn?.addEventListener('click', undoLastChange);
        redoBtn?.addEventListener('click', redoLastChange);
        exportExcelBtn?.addEventListener('click', exportUpdatedExcel);
        exportExcelBtnTop?.addEventListener('click', exportUpdatedExcel);
        exportConfluenceBtn?.addEventListener('click', exportConfluenceTable);
        exportConfluenceBtnTop?.addEventListener('click', exportConfluenceTable);
        exportConfluenceImageBtnTop?.addEventListener('click', exportConfluenceImage);

        const bindOverviewCard = (node, kind) => {{
            if (!node) {{
                return;
            }}
            node.addEventListener('click', () => openOverviewModal(kind));
            node.addEventListener('keydown', (ev) => {{
                if (ev.key === 'Enter' || ev.key === ' ') {{
                    ev.preventDefault();
                    openOverviewModal(kind);
                }}
            }});
        }};
        bindOverviewCard(overviewCardTotal, 'total');
        bindOverviewCard(overviewCardDone, 'done');
        bindOverviewCard(overviewCardPartial, 'partial');
        bindOverviewCard(overviewCardRemaining, 'remaining');
        overviewModalClose?.addEventListener('click', closeOverviewModal);
        overviewModalBackdrop?.addEventListener('click', (ev) => {{
            if (ev.target === overviewModalBackdrop) {{
                closeOverviewModal();
            }}
        }});
        restoreFromMasterBtn?.addEventListener('click', () => {{
            restoreFromMasterCopy();
        }});
        saveToMasterBtn?.addEventListener('click', () => {{
            saveToMasterCopy();
        }});
        window.__planSuggest = suggestPlanForAllServers;
        planSuggestBtn?.addEventListener('click', suggestPlanForAllServers);
        removeCountry?.addEventListener('change', refreshRemoveServerSelect);
        removeServerBtn?.addEventListener('click', removeSelectedServerRow);
        noteModalCancel?.addEventListener('click', closeNoteModal);
        noteModalBackdrop?.addEventListener('click', (ev) => {{
            if (ev.target === noteModalBackdrop) {{
                closeNoteModal();
            }}
        }});
        noteModalSave?.addEventListener('click', () => {{
            if (!activeNoteRow || !noteModalInput) {{
                closeNoteModal();
                return;
            }}
            const rowRef = activeNoteRow;
            const nextNote = String(noteModalInput.value || '').trim();
            const nextRhelOk = Boolean(noteRhelOk?.checked);
            const nextImeOk = Boolean(noteImeOk?.checked);
            const nextAppTestOk = Boolean(noteAppTestOk?.checked);
            const nextRegressionTestOk = Boolean(noteRegressionTestOk?.checked);
            const nextIpSwapOk = Boolean(noteIpSwapOk?.checked);
            const nextIpSwapStartDate = String(noteIpSwapStartDate?.value || '').trim();
            const nextIpSwapEndDate = String(noteIpSwapEndDate?.value || noteIpSwapStartDate?.value || '').trim();
            const nextIpSwapFrom = String(noteIpSwapFrom?.value || '').trim();
            const nextIpSwapTo = String(noteIpSwapTo?.value || '').trim();
            const nextPatchOk = Boolean(notePatchOk?.checked);
            const nextPatchNumber = String(notePatchNumber?.value || '').trim();
            const prevNote = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '').trim();
            const prevRhelOk = Boolean(rowRef.rhelOk);
            const prevImeOk = Boolean(rowRef.imeOk);
            const prevAppTestOk = Boolean(rowRef.appTestOk);
            const prevRegressionTestOk = Boolean(rowRef.regressionTestOk);
            const prevIpSwapOk = Boolean(rowRef.ipSwapOk);
            const prevIpSwapStartDate = String(rowRef.ipSwapStartDate || '').trim();
            const prevIpSwapEndDate = String(rowRef.ipSwapEndDate || rowRef.ipSwapStartDate || '').trim();
            const prevIpSwapFrom = String(rowRef.ipSwapFrom || '').trim();
            const prevIpSwapTo = String(rowRef.ipSwapTo || '').trim();
            const prevPatchOk = Boolean(rowRef.patchOk);
            const prevPatchNumber = String(rowRef.patchNumber || '').trim();
            if (
                nextNote === prevNote
                && nextRhelOk === prevRhelOk
                && nextImeOk === prevImeOk
                && nextAppTestOk === prevAppTestOk
                && nextRegressionTestOk === prevRegressionTestOk
                && nextIpSwapOk === prevIpSwapOk
                && nextIpSwapStartDate === prevIpSwapStartDate
                && nextIpSwapEndDate === prevIpSwapEndDate
                && nextIpSwapFrom === prevIpSwapFrom
                && nextIpSwapTo === prevIpSwapTo
                && nextPatchOk === prevPatchOk
                && nextPatchNumber === prevPatchNumber
            ) {{
                closeNoteModal();
                return;
            }}
            pushUndoSnapshot();
            rowRef.notes = nextNote;
            rowRef.notesProd = nextNote;
            rowRef.notesTest = nextNote;
            rowRef.rhelOk = nextRhelOk;
            rowRef.imeOk = nextImeOk;
            rowRef.appTestOk = nextAppTestOk;
            rowRef.regressionTestOk = nextRegressionTestOk;
            rowRef.ipSwapOk = nextIpSwapOk;
            rowRef.ipSwapStartDate = nextIpSwapStartDate;
            rowRef.ipSwapEndDate = nextIpSwapEndDate;
            rowRef.ipSwapFrom = nextIpSwapFrom;
            rowRef.ipSwapTo = nextIpSwapTo;
            rowRef.patchOk = nextPatchOk;
            rowRef.patchNumber = nextPatchNumber;
            const checksDone = [rowRef.rhelOk, rowRef.imeOk, rowRef.appTestOk, rowRef.regressionTestOk, rowRef.ipSwapOk, rowRef.patchOk].filter(Boolean).length;
            rowRef.status = checksDone === 6 ? 'done' : (checksDone > 0 ? 'partial' : 'none');
            if (typeof activeNoteBadgeUpdater === 'function') {{
                activeNoteBadgeUpdater();
            }}
            closeNoteModal();
            statusText.classList.remove('error');
            statusText.textContent = 'Server info saved.';
            renderChart();
        }});
        noteModalClear?.addEventListener('click', () => {{
            if (!activeNoteRow || !noteModalInput) {{
                closeNoteModal();
                return;
            }}
            const rowRef = activeNoteRow;
            const prevNote = String(rowRef.notes || rowRef.notesProd || rowRef.notesTest || '').trim();
            const hasSavedState = prevNote
                || Boolean(rowRef.rhelOk)
                || Boolean(rowRef.imeOk)
                || Boolean(rowRef.appTestOk)
                || Boolean(rowRef.regressionTestOk)
                || Boolean(rowRef.ipSwapOk)
                || String(rowRef.ipSwapStartDate || '').trim()
                || String(rowRef.ipSwapEndDate || '').trim()
                || String(rowRef.ipSwapFrom || '').trim()
                || String(rowRef.ipSwapTo || '').trim()
                || Boolean(rowRef.patchOk)
                || String(rowRef.patchNumber || '').trim();
            if (!hasSavedState) {{
                closeNoteModal();
                return;
            }}
            pushUndoSnapshot();
            rowRef.notesProd = '';
            rowRef.notesTest = '';
            rowRef.notes = '';
            rowRef.rhelOk = false;
            rowRef.imeOk = false;
            rowRef.appTestOk = false;
            rowRef.regressionTestOk = false;
            rowRef.ipSwapOk = false;
            rowRef.ipSwapStartDate = '';
            rowRef.ipSwapEndDate = '';
            rowRef.ipSwapFrom = '';
            rowRef.ipSwapTo = '';
            rowRef.patchOk = false;
            rowRef.patchNumber = '';
            rowRef.status = 'none';
            noteModalInput.value = '';
            if (noteRhelOk) {{
                noteRhelOk.checked = false;
            }}
            if (noteImeOk) {{
                noteImeOk.checked = false;
            }}
            if (noteAppTestOk) {{
                noteAppTestOk.checked = false;
            }}
            if (noteRegressionTestOk) {{
                noteRegressionTestOk.checked = false;
            }}
            if (noteIpSwapOk) {{
                noteIpSwapOk.checked = false;
            }}
            if (noteIpSwapStartDate) {{
                noteIpSwapStartDate.value = '';
            }}
            if (noteIpSwapEndDate) {{
                noteIpSwapEndDate.value = '';
            }}
            if (noteIpSwapFrom) {{
                noteIpSwapFrom.value = '';
            }}
            if (noteIpSwapTo) {{
                noteIpSwapTo.value = '';
            }}
            if (notePatchOk) {{
                notePatchOk.checked = false;
            }}
            if (notePatchNumber) {{
                notePatchNumber.value = '';
            }}
            if (typeof activeNoteBadgeUpdater === 'function') {{
                activeNoteBadgeUpdater();
            }}
            closeNoteModal();
            statusText.classList.remove('error');
            statusText.textContent = 'Note cleared.';
            renderChart();
        }});
        toggleFullscreenBtn?.addEventListener('click', () => {{
            toggleFullscreenView().catch(() => {{
                statusText.classList.add('error');
                statusText.textContent = 'Could not enter fullscreen in this view.';
            }});
        }});
        document.addEventListener('fullscreenchange', updateFullscreenButtonLabel);
        document.addEventListener('keydown', (ev) => {{
            if (noteModalBackdrop?.classList.contains('open')) {{
                if (ev.key === 'Escape') {{
                    ev.preventDefault();
                    closeNoteModal();
                    return;
                }}
                if ((ev.ctrlKey || ev.metaKey) && ev.key.toLowerCase() === 'enter') {{
                    ev.preventDefault();
                    noteModalSave?.click();
                    return;
                }}
            }}
            if (overviewModalBackdrop?.classList.contains('open') && ev.key === 'Escape') {{
                ev.preventDefault();
                closeOverviewModal();
                return;
            }}

            const key = ev.key.toLowerCase();
            const hasCmd = ev.ctrlKey || ev.metaKey;
            if (hasCmd && key === 'z' && ev.shiftKey) {{
                const tag = String(document.activeElement?.tagName || '').toLowerCase();
                if (tag !== 'input' && tag !== 'textarea' && !document.activeElement?.isContentEditable) {{
                    ev.preventDefault();
                    redoLastChange();
                }}
                return;
            }}

            if (hasCmd && key === 'y') {{
                const tag = String(document.activeElement?.tagName || '').toLowerCase();
                if (tag !== 'input' && tag !== 'textarea' && !document.activeElement?.isContentEditable) {{
                    ev.preventDefault();
                    redoLastChange();
                }}
                return;
            }}

            if (hasCmd && key === 'z') {{
                const tag = String(document.activeElement?.tagName || '').toLowerCase();
                if (tag !== 'input' && tag !== 'textarea' && !document.activeElement?.isContentEditable) {{
                    ev.preventDefault();
                    undoLastChange();
                }}
            }}
        }});

        updateActiveButton();
        refreshCountrySelect();
        refreshDragFilterOptions();
        updateFullscreenButtonLabel();
        updateUndoButton();
        updateRedoButton();
        const restored = loadSavedState();
        if (restored) {{
            statusText.classList.remove('error');
            statusText.textContent = 'Loaded previously saved changes.';
        }}
        renderChart();
    </script>
</body>
</html>
'''


def verify_drag_editor_integrity(html_text: str, view_name: str) -> None:
    missing = [snippet for snippet in DRAG_EDITOR_REQUIRED_SNIPPETS if snippet not in html_text]
    if missing:
        raise RuntimeError(
            f"{view_name}: drag editor is missing required parts: " + ", ".join(missing)
        )


def verify_against_backup_markers(html_text: str) -> None:
    if not BACKUP_STATUS_HTML.exists():
        return
    backup_text = BACKUP_STATUS_HTML.read_text(encoding="utf-8")
    baseline_missing = [snippet for snippet in DRAG_EDITOR_REQUIRED_SNIPPETS if snippet not in backup_text]
    if baseline_missing:
        # Older backups can be stale; keep generation unblocked and validate current output instead.
        return
    generated_missing = [snippet for snippet in DRAG_EDITOR_REQUIRED_SNIPPETS if snippet not in html_text]
    if generated_missing:
        raise RuntimeError(
            "Generated status view does not match the backup drag editor markers. Missing: "
            + ", ".join(generated_missing)
        )


def main():
    rows = parse_rows()
    countries = country_order(rows)

    mermaid_plain, frame_start, last_date = build_mermaid(rows, countries, title="Upgrade Plan 2026")

    generated_md = (
        "# Gantt chart from Tidplan.xlsx\n\n"
        "Automatically generated from the planning sheet.\n\n"
        "```mermaid\n"
        + mermaid_plain
        + "```\n"
    )

    status_md = (
        "# Gantt chart by country\n\n"
        "All countries in one view with filtering in the HTML view.\n\n"
        "```mermaid\n"
        + mermaid_plain
        + "```\n\n"
        "HTML view (production): tidplan_gantt_status_per_land.html\n"
        "HTML view (planning): tidplan_gantt_kalkyl.html\n"
    )

    html_prod = build_html(rows, countries, frame_start, last_date, planner_mode=False)
    html_calc = build_html(rows, countries, frame_start, last_date, planner_mode=True)

    verify_drag_editor_integrity(html_prod, "status view")
    verify_drag_editor_integrity(html_calc, "planning view")
    verify_against_backup_markers(html_prod)

    OUT_MD.write_text(generated_md, encoding="utf-8")
    OUT_STATUS_MD.write_text(status_md, encoding="utf-8")
    OUT_STATUS_HTML.write_text(html_prod, encoding="utf-8")
    OUT_CALC_HTML.write_text(html_calc, encoding="utf-8")
    OUT_INDEX_HTML.write_text(html_prod, encoding="utf-8")

    print("Updated files:")
    print("- tidplan_gantt_generated.md")
    print("- tidplan_gantt_status_per_land.md")
    print("- tidplan_gantt_status_per_land.html")
    print("- tidplan_gantt_kalkyl.html")
    print("- index.html")
    print("- Drag editor integrity check: OK")
    print(f"Date range: {iso(frame_start)} to {iso(last_date)}")
    print(f"Countries: {', '.join(countries)}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()

